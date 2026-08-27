"""
Non-SEBI Creator Cumulative Ledger & Manager Submission Engine
==============================================================
Tracks and manages all non-SEBI-registered creators using Telegram integration
across daily settlement audits until 5 September 2026.

Guarantees:
  1. Strict deduplication by creatorId (one unique row per creator).
  2. Running cumulative volume and violation day counter across audit dates.
  3. Automatic persistence in:
     - data/non_sebi_creators_ledger.json
     - data/non_sebi_creators_cumulative.xlsx (multi-sheet formatted workbook)
     - data/non_sebi_creators_cumulative.csv
     - reports/Manager_Submission_Non_SEBI_Report.html
     - reports/Manager_Submission_Non_SEBI_Report.pdf
"""

import os
import re
import json
import datetime
from typing import Dict, List, Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = HERE
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
REPORTS_DIR = os.path.join(PROJECT_ROOT, "reports")

LEDGER_JSON_FILE = os.path.join(DATA_DIR, "non_sebi_creators_ledger.json")
CUMULATIVE_EXCEL_FILE = os.path.join(DATA_DIR, "non_sebi_creators_cumulative.xlsx")
CUMULATIVE_CSV_FILE = os.path.join(DATA_DIR, "non_sebi_creators_cumulative.csv")
MANAGER_HTML_REPORT = os.path.join(REPORTS_DIR, "Manager_Submission_Non_SEBI_Report.html")
MANAGER_PDF_REPORT = os.path.join(REPORTS_DIR, "Manager_Submission_Non_SEBI_Report.pdf")

TARGET_SUBMISSION_DATE = "2026-09-05"


def normalize_id(cid: Any) -> str:
    """Normalize Creator ID to lowercase trimmed string."""
    if not cid:
        return ""
    cid_str = str(cid).strip().lower()
    m = re.search(r"([0-9a-f]{24})", cid_str)
    return m.group(1) if m else cid_str


def load_ledger() -> Dict[str, Any]:
    """Load existing non-SEBI cumulative ledger from JSON."""
    os.makedirs(DATA_DIR, exist_ok=True)
    if os.path.exists(LEDGER_JSON_FILE):
        try:
            with open(LEDGER_JSON_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass

    return {
        "metadata": {
            "title": "Cosmofeed Non-SEBI Creator Cumulative Ledger",
            "auditPeriodStart": "2026-08-27",
            "auditPeriodEnd": TARGET_SUBMISSION_DATE,
            "targetSubmissionDate": TARGET_SUBMISSION_DATE,
            "lastUpdated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "totalAuditRuns": 0,
            "auditDates": []
        },
        "creators": {},
        "dailySnapshots": {}
    }


def save_ledger(ledger: Dict[str, Any]) -> None:
    """Save non-SEBI cumulative ledger to JSON."""
    os.makedirs(DATA_DIR, exist_ok=True)
    ledger["metadata"]["lastUpdated"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LEDGER_JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(ledger, f, indent=2, ensure_ascii=False)


def record_daily_settlements(creators: List[Dict[str, Any]], audit_date: Optional[str] = None) -> Dict[str, Any]:
    """
    Process daily settlements, filter non-SEBI creators using Telegram integration,
    and deduplicate/upsert them into the persistent ledger.
    """
    if not audit_date:
        audit_date = datetime.date.today().strftime("%Y-%m-%d")

    ledger = load_ledger()
    creators_map = ledger.get("creators", {})
    daily_snapshots = ledger.get("dailySnapshots", {})
    audit_dates = ledger["metadata"].get("auditDates", [])

    if audit_date not in audit_dates:
        audit_dates.append(audit_date)
        audit_dates.sort()
        ledger["metadata"]["auditDates"] = audit_dates
    ledger["metadata"]["totalAuditRuns"] = len(audit_dates)

    today_flagged = []

    for r in creators:
        # Check if creator is eligible: Telegram integration AND non-SEBI registered
        is_telegram = bool(r.get("telegramIntegration") and r.get("telegramEligible"))
        is_non_sebi = (r.get("sebiRegisteredNo") == "No") or (r.get("sebiVerificationStatus") in ["Not Verified", "Manual Review Required"])

        if not (is_telegram and is_non_sebi):
            continue

        cid = normalize_id(r.get("creatorId") or r.get("settlementId"))
        if not cid:
            continue

        try:
            payout = float(r.get("payoutAmount") or 0.0)
        except (ValueError, TypeError):
            payout = 0.0

        today_flagged.append(cid)

        # Upsert into creators_map with strict deduplication
        if cid not in creators_map:
            creators_map[cid] = {
                "creatorId": cid,
                "username": r.get("username") or "N/A",
                "creatorName": r.get("displayName") or r.get("creatorName") or r.get("username") or "Creator",
                "email": r.get("email") or "N/A",
                "phone": r.get("phone") or "N/A",
                "onboardedBy": r.get("onboardedBy") or "N/A",
                "onboardingVertical": r.get("onboardingVertical") or r.get("vertical") or "N/A",
                "telegramProductId": r.get("telegramProductId") or "N/A",
                "telegramProductLink": r.get("telegramProductLink") or "N/A",
                "telegramProductIds": [r.get("telegramProductId")] if r.get("telegramProductId") else [],
                "telegramProductLinks": [r.get("telegramProductLink")] if r.get("telegramProductLink") else [],
                "sebiVerificationStatus": "Not Verified",
                "sebiRegisteredYes": "—",
                "sebiRegisteredNo": "No",
                "reviewStatus": "Manual Review Required",
                "complianceHoldStatus": "HOLD - RELEASE RESTRICTED",
                "firstSeenDate": audit_date,
                "lastSeenDate": audit_date,
                "datesObserved": [audit_date],
                "daysFlaggedCount": 1,
                "latestPayoutAmount": payout,
                "cumulativePayoutVolume": payout,
                "maxDailyPayout": payout,
                "settlementHistory": [
                    {
                        "auditDate": audit_date,
                        "payoutAmount": payout,
                        "telegramProductId": r.get("telegramProductId") or "N/A",
                        "telegramProductLink": r.get("telegramProductLink") or "N/A"
                    }
                ]
            }
        else:
            rec = creators_map[cid]
            rec["lastSeenDate"] = audit_date
            if audit_date not in rec["datesObserved"]:
                rec["datesObserved"].append(audit_date)
                rec["datesObserved"].sort()
                rec["daysFlaggedCount"] = len(rec["datesObserved"])
                rec["cumulativePayoutVolume"] = round(rec["cumulativePayoutVolume"] + payout, 2)
            
            rec["latestPayoutAmount"] = payout
            rec["maxDailyPayout"] = max(rec.get("maxDailyPayout", 0.0), payout)
            
            # Keep product list fresh
            p_id = r.get("telegramProductId")
            if p_id and p_id not in rec["telegramProductIds"]:
                rec["telegramProductIds"].append(p_id)
            p_link = r.get("telegramProductLink")
            if p_link and p_link not in rec["telegramProductLinks"]:
                rec["telegramProductLinks"].append(p_link)
            
            # History log (one record per date)
            existing_hist_dates = [h["auditDate"] for h in rec["settlementHistory"]]
            if audit_date not in existing_hist_dates:
                rec["settlementHistory"].append({
                    "auditDate": audit_date,
                    "payoutAmount": payout,
                    "telegramProductId": r.get("telegramProductId") or "N/A",
                    "telegramProductLink": r.get("telegramProductLink") or "N/A"
                })

    # Save daily snapshot
    daily_snapshots[audit_date] = {
        "auditDate": audit_date,
        "flaggedCount": len(today_flagged),
        "totalDayVolume": sum(creators_map[cid]["latestPayoutAmount"] for cid in today_flagged if cid in creators_map),
        "creatorIds": today_flagged
    }

    ledger["creators"] = creators_map
    ledger["dailySnapshots"] = daily_snapshots
    save_ledger(ledger)

    # Export to Excel, CSV, HTML, and PDF
    export_cumulative_excel(ledger)
    export_cumulative_csv(ledger)
    generate_manager_html_report(ledger)
    generate_manager_pdf_report(ledger)

    return ledger


def get_sorted_non_sebi_list(ledger: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Return deduplicated list of non-SEBI creators sorted by cumulative volume descending."""
    creators_map = ledger.get("creators", {})
    records = list(creators_map.values())
    records.sort(key=lambda x: (x.get("cumulativePayoutVolume", 0.0), x.get("latestPayoutAmount", 0.0)), reverse=True)
    return records


def export_cumulative_excel(ledger: Optional[Dict[str, Any]] = None) -> str:
    """Generate high-quality multi-sheet Excel report for manager submission."""
    if ledger is None:
        ledger = load_ledger()

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    records = get_sorted_non_sebi_list(ledger)
    meta = ledger.get("metadata", {})
    snapshots = ledger.get("dailySnapshots", {})

    total_cumulative_vol = sum(r.get("cumulativePayoutVolume", 0.0) for r in records)
    total_creators_cnt = len(records)
    audit_dates = meta.get("auditDates", [])

    # Color Tokens
    NAVY_FILL = PatternFill("solid", fgColor="1E293B")
    BLUE_FILL = PatternFill("solid", fgColor="3B82F6")
    ROSE_FILL = PatternFill("solid", fgColor="E11D48")
    LIGHT_GRAY_FILL = PatternFill("solid", fgColor="F8FAFC")
    ZEBRA_FILL = PatternFill("solid", fgColor="F1F5F9")
    HOLD_FILL = PatternFill("solid", fgColor="FFE4E6")

    FONT_TITLE = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    FONT_HEADER = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    FONT_BOLD = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    FONT_NORMAL = Font(name="Segoe UI", size=10, color="334155")
    FONT_ROSE_BOLD = Font(name="Segoe UI", size=10, bold=True, color="BE123C")
    FONT_KPI_NUM = Font(name="Segoe UI", size=16, bold=True, color="0F172A")
    FONT_KPI_LBL = Font(name="Segoe UI", size=9, bold=True, color="64748B")

    THIN_BORDER = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    # -------------------------------------------------------------
    # SHEET 1: EXECUTIVE SUMMARY
    # -------------------------------------------------------------
    ws_sum = wb.create_sheet(title="Executive Summary")
    ws_sum.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_sum.merge_cells("A1:G2")
    ws_sum["A1"] = "COSMOFEED COMPLIANCE AUDIT — NON-SEBI CREATOR SUBMISSION"
    ws_sum["A1"].font = FONT_TITLE
    ws_sum["A1"].fill = NAVY_FILL
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_sum["A4"] = f"Audit Window: {meta.get('auditPeriodStart')} to {TARGET_SUBMISSION_DATE} (Target Manager Submission)"
    ws_sum["A4"].font = FONT_BOLD

    ws_sum["A5"] = f"Generated On: {meta.get('lastUpdated')} | Daily Audits Logged: {len(audit_dates)}"
    ws_sum["A5"].font = FONT_NORMAL

    # KPI Summary Cards
    ws_sum["A7"] = "TOTAL NON-SEBI CREATORS"
    ws_sum["A7"].font = FONT_KPI_LBL
    ws_sum["A8"] = total_creators_cnt
    ws_sum["A8"].font = FONT_KPI_NUM

    ws_sum["C7"] = "CUMULATIVE VOLUME ON HOLD"
    ws_sum["C7"].font = FONT_KPI_LBL
    ws_sum["C8"] = f"₹{total_cumulative_vol:,.2f}"
    ws_sum["C8"].font = FONT_KPI_NUM

    ws_sum["E7"] = "AUDIT SUBMISSION STATUS"
    ws_sum["E7"].font = FONT_KPI_LBL
    ws_sum["E8"] = "Active Tracking Until 05-Sep"
    ws_sum["E8"].font = FONT_KPI_NUM

    # Daily Breakdown Table in Summary
    ws_sum["A11"] = "Daily Audit Ingestion History"
    ws_sum["A11"].font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")

    sum_headers = ["Audit Date", "Flagged Creators", "Day Settlement Volume (₹)", "Cumulative Tracking Status"]
    for col_num, h_text in enumerate(sum_headers, start=1):
        cell = ws_sum.cell(row=12, column=col_num)
        cell.value = h_text
        cell.font = FONT_HEADER
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 13
    for d_str, snap in sorted(snapshots.items()):
        ws_sum.cell(row=row_idx, column=1, value=d_str).font = FONT_NORMAL
        ws_sum.cell(row=row_idx, column=2, value=snap.get("flaggedCount", 0)).font = FONT_BOLD
        c3 = ws_sum.cell(row=row_idx, column=3, value=round(snap.get("totalDayVolume", 0.0), 2))
        c3.font = FONT_BOLD
        c3.number_format = '"₹"#,##0.00'
        ws_sum.cell(row=row_idx, column=4, value="Logged & Deduplicated").font = FONT_NORMAL
        for c in range(1, 5):
            ws_sum.cell(row=row_idx, column=c).border = THIN_BORDER
        row_idx += 1

    # -------------------------------------------------------------
    # SHEET 2: MASTER NON-SEBI REGISTRY (DEDUPLICATED)
    # -------------------------------------------------------------
    ws_reg = wb.create_sheet(title="Non-SEBI Master Registry")
    ws_reg.views.sheetView[0].showGridLines = True

    reg_headers = [
        "#",
        "Creator ID",
        "Creator Username",
        "Contact Email",
        "Contact Phone",
        "Onboarded By",
        "First Seen Date",
        "Latest Seen Date",
        "Days Flagged",
        "Latest Settlement (₹)",
        "Cumulative Volume (₹)",
        "Max Daily Payout (₹)",
        "Telegram Product ID",
        "Telegram Product Link",
        "SEBI Status",
        "Compliance Action"
    ]

    for col_num, h_text in enumerate(reg_headers, start=1):
        cell = ws_reg.cell(row=1, column=col_num)
        cell.value = h_text
        cell.font = FONT_HEADER
        cell.fill = NAVY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_reg.row_dimensions[1].height = 28

    for i, r in enumerate(records, start=1):
        row_num = i + 1
        fill = ZEBRA_FILL if i % 2 == 0 else LIGHT_GRAY_FILL

        ws_reg.cell(row=row_num, column=1, value=i).font = FONT_NORMAL
        ws_reg.cell(row=row_num, column=2, value=r.get("creatorId")).font = FONT_BOLD
        ws_reg.cell(row=row_num, column=3, value=r.get("username")).font = FONT_BOLD
        ws_reg.cell(row=row_num, column=4, value=r.get("email")).font = FONT_NORMAL
        ws_reg.cell(row=row_num, column=5, value=r.get("phone")).font = FONT_NORMAL
        ws_reg.cell(row=row_num, column=6, value=r.get("onboardedBy")).font = FONT_NORMAL
        ws_reg.cell(row=row_num, column=7, value=r.get("firstSeenDate")).font = FONT_NORMAL
        ws_reg.cell(row=row_num, column=8, value=r.get("lastSeenDate")).font = FONT_NORMAL
        ws_reg.cell(row=row_num, column=9, value=r.get("daysFlaggedCount", 1)).font = FONT_BOLD

        c_latest = ws_reg.cell(row=row_num, column=10, value=r.get("latestPayoutAmount", 0.0))
        c_latest.font = FONT_ROSE_BOLD
        c_latest.number_format = '"₹"#,##0.00'

        c_cum = ws_reg.cell(row=row_num, column=11, value=r.get("cumulativePayoutVolume", 0.0))
        c_cum.font = FONT_ROSE_BOLD
        c_cum.number_format = '"₹"#,##0.00'

        c_max = ws_reg.cell(row=row_num, column=12, value=r.get("maxDailyPayout", 0.0))
        c_max.font = FONT_NORMAL
        c_max.number_format = '"₹"#,##0.00'

        ws_reg.cell(row=row_num, column=13, value=r.get("telegramProductId")).font = FONT_NORMAL

        c_link = ws_reg.cell(row=row_num, column=14, value=r.get("telegramProductLink"))
        c_link.font = FONT_NORMAL
        if str(r.get("telegramProductLink", "")).startswith("http"):
            c_link.hyperlink = r.get("telegramProductLink")
            c_link.font = Font(name="Segoe UI", size=10, color="2563EB", underline="single")

        ws_reg.cell(row=row_num, column=15, value=r.get("sebiVerificationStatus", "Not Verified")).font = FONT_ROSE_BOLD

        c_action = ws_reg.cell(row=row_num, column=16, value=r.get("complianceHoldStatus", "HOLD"))
        c_action.font = FONT_ROSE_BOLD
        c_action.fill = HOLD_FILL

        for col in range(1, 17):
            c_cell = ws_reg.cell(row=row_num, column=col)
            c_cell.border = THIN_BORDER
            if col != 16:
                c_cell.fill = fill

    # -------------------------------------------------------------
    # Auto-fit Column Widths
    # -------------------------------------------------------------
    for ws in [ws_sum, ws_reg]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format and '"₹"' in cell.number_format:
                    val += "   "
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    os.makedirs(DATA_DIR, exist_ok=True)
    wb.save(CUMULATIVE_EXCEL_FILE)
    return CUMULATIVE_EXCEL_FILE


def export_cumulative_csv(ledger: Optional[Dict[str, Any]] = None) -> str:
    """Generate clean CSV export of deduplicated non-SEBI records."""
    if ledger is None:
        ledger = load_ledger()

    records = get_sorted_non_sebi_list(ledger)
    headers = [
        "Creator ID",
        "Username",
        "Email",
        "Phone",
        "Onboarded By",
        "First Seen Date",
        "Latest Seen Date",
        "Days Flagged",
        "Latest Settlement (INR)",
        "Cumulative Volume (INR)",
        "Max Daily Payout (INR)",
        "Telegram Product ID",
        "Telegram Product Link",
        "SEBI Verification Status",
        "Compliance Action"
    ]

    import csv
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(CUMULATIVE_CSV_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in records:
            writer.writerow([
                r.get("creatorId", ""),
                r.get("username", ""),
                r.get("email", ""),
                r.get("phone", ""),
                r.get("onboardedBy", ""),
                r.get("firstSeenDate", ""),
                r.get("lastSeenDate", ""),
                r.get("daysFlaggedCount", 1),
                r.get("latestPayoutAmount", 0.0),
                r.get("cumulativePayoutVolume", 0.0),
                r.get("maxDailyPayout", 0.0),
                r.get("telegramProductId", ""),
                r.get("telegramProductLink", ""),
                r.get("sebiVerificationStatus", "Not Verified"),
                r.get("complianceHoldStatus", "HOLD")
            ])

    return CUMULATIVE_CSV_FILE


def generate_manager_html_report(ledger: Optional[Dict[str, Any]] = None) -> str:
    """Generate executive HTML report formatted specifically for submission to the manager."""
    if ledger is None:
        ledger = load_ledger()

    records = get_sorted_non_sebi_list(ledger)
    meta = ledger.get("metadata", {})
    total_volume = sum(r.get("cumulativePayoutVolume", 0.0) for r in records)
    total_count = len(records)
    audit_dates = meta.get("auditDates", [])

    rows_html = []
    for i, r in enumerate(records, start=1):
        p_link = r.get("telegramProductLink") or "#"
        p_id = r.get("telegramProductId") or "N/A"
        rows_html.append(f"""
        <tr class="hover:bg-rose-50/50 border-b border-slate-100 transition text-xs">
          <td class="py-2.5 px-3 font-semibold text-slate-400">{i}</td>
          <td class="py-2.5 px-3 font-mono text-slate-600 bg-slate-50 rounded px-1.5 py-0.5">{r.get('creatorId')}</td>
          <td class="py-2.5 px-3 font-bold text-slate-900">{r.get('username')}</td>
          <td class="py-2.5 px-3 text-slate-600">{r.get('email')}</td>
          <td class="py-2.5 px-3 text-slate-600">{r.get('phone')}</td>
          <td class="py-2.5 px-3 text-slate-700 font-medium">{r.get('onboardedBy')}</td>
          <td class="py-2.5 px-3 text-center text-slate-600">{r.get('firstSeenDate')}</td>
          <td class="py-2.5 px-3 text-center font-bold text-slate-800">{r.get('lastSeenDate')}</td>
          <td class="py-2.5 px-3 text-center font-extrabold text-blue-600">{r.get('daysFlaggedCount', 1)}</td>
          <td class="py-2.5 px-3 text-right font-extrabold text-slate-900 font-mono">₹{r.get('latestPayoutAmount', 0.0):,.2f}</td>
          <td class="py-2.5 px-3 text-right font-extrabold text-rose-600 font-mono">₹{r.get('cumulativePayoutVolume', 0.0):,.2f}</td>
          <td class="py-2.5 px-3 font-mono text-[11px]">
            <a href="{p_link}" target="_blank" class="text-blue-600 hover:underline">vig/{p_id[:10]}..</a>
          </td>
          <td class="py-2.5 px-3 text-center">
            <span class="px-2 py-0.5 rounded-full text-[10px] font-extrabold bg-rose-100 text-rose-800 border border-rose-200">HOLD</span>
          </td>
        </tr>
        """)

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Manager Submission — Non-SEBI Creator Cumulative Audit Report (Until 5 Sep)</title>
<script src="https://cdn.tailwindcss.com"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=JetBrains+Mono:wght@500;600&display=swap" rel="stylesheet">
<style>
  body {{ font-family: 'Inter', sans-serif; background: #f8fafc; color: #0f172a; }}
  @media print {{
    .no-print {{ display: none !important; }}
    body {{ background: #fff; }}
  }}
</style>
</head>
<body class="p-6 max-w-7xl mx-auto">
  <!-- Header Block -->
  <div class="bg-white rounded-2xl p-6 shadow-sm border border-slate-200 mb-6">
    <div class="flex flex-col md:flex-row md:items-center justify-between gap-4 border-b border-slate-100 pb-5">
      <div>
        <div class="inline-flex items-center gap-2 px-3 py-1 rounded-full text-xs font-bold bg-rose-50 text-rose-700 border border-rose-200 mb-2">
          <span>🚨 EXECUTIVE MANAGER SUBMISSION REPORT</span>
          <span>·</span>
          <span>Target Deadline: {TARGET_SUBMISSION_DATE}</span>
        </div>
        <h1 class="text-2xl font-extrabold text-slate-900 tracking-tight">Cosmofeed Payout Audit: Non-SEBI Creator Cumulative Registry</h1>
        <p class="text-xs text-slate-500 mt-1">Audit Tracking Window: <b>{meta.get('auditPeriodStart')}</b> to <b>{TARGET_SUBMISSION_DATE}</b> | Generated on: <b>{meta.get('lastUpdated')}</b></p>
      </div>
      <div class="flex items-center gap-2 no-print">
        <button onclick="window.print()" class="px-4 py-2 bg-slate-900 hover:bg-slate-800 text-white text-xs font-bold rounded-xl shadow transition">
          Print / Save PDF
        </button>
        <a href="/api/non-sebi/download-excel" class="px-4 py-2 bg-emerald-600 hover:bg-emerald-700 text-white text-xs font-bold rounded-xl shadow transition">
          Download Master Excel (.xlsx)
        </a>
      </div>
    </div>

    <!-- KPI Summary Cards -->
    <div class="grid grid-cols-2 md:grid-cols-4 gap-4 mt-6">
      <div class="p-4 rounded-xl bg-slate-50 border border-slate-200/80">
        <div class="text-2xl font-extrabold text-slate-900">{total_count}</div>
        <div class="text-[11px] font-bold text-slate-500 uppercase mt-1">Unique Flagged Creators</div>
        <div class="text-[10px] text-slate-400 mt-0.5">Strictly Deduplicated</div>
      </div>
      <div class="p-4 rounded-xl bg-rose-50 border border-rose-200/80">
        <div class="text-2xl font-extrabold text-rose-600">₹{total_volume:,.2f}</div>
        <div class="text-[11px] font-bold text-rose-700 uppercase mt-1">Cumulative Volume on Hold</div>
        <div class="text-[10px] text-rose-500 mt-0.5">All settlement batches</div>
      </div>
      <div class="p-4 rounded-xl bg-blue-50 border border-blue-200/80">
        <div class="text-2xl font-extrabold text-blue-700">{len(audit_dates)}</div>
        <div class="text-[11px] font-bold text-blue-800 uppercase mt-1">Daily Audits Logged</div>
        <div class="text-[10px] text-blue-600 mt-0.5">{", ".join(audit_dates[-3:]) if audit_dates else 'None'}</div>
      </div>
      <div class="p-4 rounded-xl bg-amber-50 border border-amber-200/80">
        <div class="text-2xl font-extrabold text-amber-700">100% HOLD</div>
        <div class="text-[11px] font-bold text-amber-800 uppercase mt-1">Mandatory Policy</div>
        <div class="text-[10px] text-amber-600 mt-0.5">Release requires SEBI verification</div>
      </div>
    </div>

    <!-- Manager Summary Notice -->
    <div class="mt-6 p-4 rounded-xl bg-slate-900 text-slate-200 text-xs leading-relaxed">
      <p class="font-semibold text-white mb-1">Executive Compliance Brief for Management:</p>
      This ledger contains all creators utilizing Cosmofeed's Telegram integration (<code class="text-rose-300">vig/productId</code>) with payout volume &ge; ₹1,000 who are <b>NOT registered with SEBI</b> in the organization's master registry. 
      Under regulatory standards, all settlements for these creators are placed on mandatory <b>HOLD</b> pending verified registration documents. This report is maintained daily until the final review date of <b>{TARGET_SUBMISSION_DATE}</b>.
    </div>
  </div>

  <!-- Master Deduplicated Registry Table -->
  <div class="bg-white rounded-2xl shadow-sm border border-slate-200 overflow-hidden mb-8">
    <div class="p-4 border-b border-slate-100 flex items-center justify-between">
      <h3 class="font-bold text-slate-800 text-sm">Deduplicated Non-SEBI Creator Ledger ({total_count} Creators)</h3>
      <span class="text-xs text-slate-500">Sorted by Cumulative Volume (Highest &rarr; Lowest)</span>
    </div>
    <div class="overflow-x-auto">
      <table class="w-full text-left">
        <thead>
          <tr class="bg-slate-100/90 text-slate-700 text-[11px] font-bold uppercase tracking-wider border-b border-slate-200">
            <th class="py-3 px-3 w-10">#</th>
            <th class="py-3 px-3">Creator ID</th>
            <th class="py-3 px-3">Username</th>
            <th class="py-3 px-3">Email</th>
            <th class="py-3 px-3">Phone</th>
            <th class="py-3 px-3">Onboarder</th>
            <th class="py-3 px-3 text-center">First Seen</th>
            <th class="py-3 px-3 text-center">Latest Seen</th>
            <th class="py-3 px-3 text-center">Days</th>
            <th class="py-3 px-3 text-right">Latest Payout</th>
            <th class="py-3 px-3 text-right">Cumulative Vol</th>
            <th class="py-3 px-3">Product Link</th>
            <th class="py-3 px-3 text-center">Action</th>
          </tr>
        </thead>
        <tbody class="divide-y divide-slate-100">
          {"".join(rows_html)}
        </tbody>
      </table>
    </div>
  </div>

  <!-- Manager Sign-off Block -->
  <div class="bg-white rounded-2xl p-6 shadow-sm border border-slate-200 text-xs">
    <h4 class="font-bold text-slate-900 mb-4 uppercase tracking-wider text-[11px]">Audit Sign-off & Acknowledgement</h4>
    <div class="grid grid-cols-1 md:grid-cols-3 gap-6 pt-2">
      <div class="border-t border-slate-300 pt-2">
        <p class="font-bold text-slate-700">Prepared By:</p>
        <p class="text-slate-500 mt-1">Cosmofeed Automated Payout Audit System</p>
        <p class="text-slate-400 text-[10px] mt-0.5">Signature: ______________________</p>
      </div>
      <div class="border-t border-slate-300 pt-2">
        <p class="font-bold text-slate-700">Reviewed By (Compliance Team):</p>
        <p class="text-slate-500 mt-1">SEBI Compliance & Risk Operations</p>
        <p class="text-slate-400 text-[10px] mt-0.5">Signature: ______________________</p>
      </div>
      <div class="border-t border-slate-300 pt-2">
        <p class="font-bold text-slate-700">Approved By (Manager):</p>
        <p class="text-slate-500 mt-1">Finance & Settlements Manager</p>
        <p class="text-slate-400 text-[10px] mt-0.5">Date: ______ / ______ / 2026</p>
      </div>
    </div>
  </div>
</body>
</html>
"""

    os.makedirs(REPORTS_DIR, exist_ok=True)
    with open(MANAGER_HTML_REPORT, "w", encoding="utf-8") as f:
        f.write(html_content)

    return MANAGER_HTML_REPORT


def generate_manager_pdf_report(ledger: Optional[Dict[str, Any]] = None) -> str:
    """Generate formal executive Manager Submission PDF report."""
    if ledger is None:
        ledger = load_ledger()

    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.pdfgen import canvas

    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.draw_footer(num_pages)
                super().showPage()
            super().save()

        def draw_footer(self, page_count):
            self.saveState()
            self.setFont("Helvetica", 8)
            self.setFillColor(colors.HexColor("#64748B"))
            text = f"Cosmofeed Compliance Audit — Non-SEBI Creator Submission | Page {self._pageNumber} of {page_count}"
            self.drawRightString(762, 18, text)
            self.drawString(30, 18, f"Confidential & Privileged — Target Deadline: {TARGET_SUBMISSION_DATE}")
            self.restoreState()

    doc = SimpleDocTemplate(
        MANAGER_PDF_REPORT,
        pagesize=landscape(letter),
        leftMargin=30,
        rightMargin=30,
        topMargin=30,
        bottomMargin=35
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=14, leading=17, textColor=colors.HexColor('#0F172A'))
    sub_style = ParagraphStyle('SubStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=8.5, leading=11, textColor=colors.HexColor('#475569'))
    th_style = ParagraphStyle('THStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white, alignment=1)
    td_style = ParagraphStyle('TDStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#1E293B'))
    td_bold = ParagraphStyle('TDBold', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#0F172A'))
    td_rose = ParagraphStyle('TDRose', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#BE123C'), alignment=2)
    td_right = ParagraphStyle('TDRight', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#0F172A'), alignment=2)

    records = get_sorted_non_sebi_list(ledger)
    meta = ledger.get("metadata", {})
    total_volume = sum(r.get("cumulativePayoutVolume", 0.0) for r in records)
    total_count = len(records)

    story = []

    # Title & Subtitle
    story.append(Paragraph("COSMOFEED COMPLIANCE AUDIT — NON-SEBI CREATOR SUBMISSION", title_style))
    story.append(Spacer(1, 3))
    story.append(Paragraph(
        f"Audit Period: <b>{meta.get('auditPeriodStart')}</b> to <b>{TARGET_SUBMISSION_DATE}</b> | "
        f"Generated: <b>{meta.get('lastUpdated')}</b> | "
        f"Unique Flagged Creators: <b>{total_count}</b> | "
        f"Total Cumulative Volume Held: <b>₹{total_volume:,.2f}</b>",
        sub_style
    ))
    story.append(Spacer(1, 10))

    # Executive KPI Summary Table
    kpi_data = [
        [
            Paragraph(f"<b>{total_count}</b><br/><font size=7 color='#64748B'>UNIQUE NON-SEBI CREATORS</font>", td_style),
            Paragraph(f"<b>₹{total_volume:,.2f}</b><br/><font size=7 color='#64748B'>CUMULATIVE VOLUME ON HOLD</font>", td_style),
            Paragraph(f"<b>{len(meta.get('auditDates', []))}</b><br/><font size=7 color='#64748B'>DAILY AUDITS LOGGED</font>", td_style),
            Paragraph(f"<b>HOLD RESTRICTED</b><br/><font size=7 color='#64748B'>SEBI COMPLIANCE MANDATE</font>", td_style),
        ]
    ]
    t_kpi = Table(kpi_data, colWidths=[185, 185, 185, 185])
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    story.append(t_kpi)
    story.append(Spacer(1, 12))

    # Table of Creators
    col_widths = [24, 115, 100, 75, 55, 55, 30, 70, 75, 75, 48]
    table_data = [[
        Paragraph("#", th_style),
        Paragraph("Creator ID", th_style),
        Paragraph("Username", th_style),
        Paragraph("Onboarder", th_style),
        Paragraph("First Seen", th_style),
        Paragraph("Last Seen", th_style),
        Paragraph("Days", th_style),
        Paragraph("Latest Payout", th_style),
        Paragraph("Cumulative Vol", th_style),
        Paragraph("Product ID", th_style),
        Paragraph("Action", th_style)
    ]]

    # Include top 80 creators in PDF or all if feasible
    pdf_records = records[:80]
    for i, r in enumerate(pdf_records, start=1):
        table_data.append([
            Paragraph(str(i), td_style),
            Paragraph(r.get("creatorId", ""), td_style),
            Paragraph(r.get("username", "")[:18], td_bold),
            Paragraph(r.get("onboardedBy", "")[:14], td_style),
            Paragraph(r.get("firstSeenDate", ""), td_style),
            Paragraph(r.get("lastSeenDate", ""), td_style),
            Paragraph(str(r.get("daysFlaggedCount", 1)), td_style),
            Paragraph(f"₹{r.get('latestPayoutAmount', 0.0):,.2f}", td_right),
            Paragraph(f"₹{r.get('cumulativePayoutVolume', 0.0):,.2f}", td_rose),
            Paragraph(r.get("telegramProductId", "")[:14], td_style),
            Paragraph("HOLD", ParagraphStyle('Hold', parent=td_style, fontName='Helvetica-Bold', textColor=colors.HexColor('#BE123C'), alignment=1))
        ])

    t_main = Table(table_data, colWidths=col_widths, repeatRows=1)
    t_main.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    story.append(t_main)
    story.append(Spacer(1, 15))

    # Sign-off block
    sign_data = [
        [
            Paragraph("<b>Prepared By:</b><br/>Cosmofeed Automated Payout Audit System", td_style),
            Paragraph("<b>Reviewed By:</b><br/>SEBI Compliance & Risk Team", td_style),
            Paragraph("<b>Manager Approval:</b><br/>Finance & Settlements Operations", td_style)
        ],
        [
            Paragraph("Signature: ________________________", td_style),
            Paragraph("Signature: ________________________", td_style),
            Paragraph("Date: _____ / _____ / 2026", td_style)
        ]
    ]
    t_sign = Table(sign_data, colWidths=[245, 245, 245])
    t_sign.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F1F5F9')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t_sign)

    doc.build(story, canvasmaker=NumberedCanvas)
    return MANAGER_PDF_REPORT


if __name__ == "__main__":
    # Test initialization using reports/data.json if present
    data_path = os.path.join(REPORTS_DIR, "data.json")
    if os.path.exists(data_path):
        with open(data_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        creators = data.get("creators", [])
        audit_date = data.get("reviewDate") or "2026-08-27"
        print(f"Testing non_sebi_manager with {len(creators)} creators on {audit_date}...")
        res = record_daily_settlements(creators, audit_date)
        print(f"Successfully processed {len(res.get('creators', {}))} non-SEBI creators!")
        print(f"Excel: {CUMULATIVE_EXCEL_FILE}")
        print(f"HTML: {MANAGER_HTML_REPORT}")
        print(f"PDF: {MANAGER_PDF_REPORT}")

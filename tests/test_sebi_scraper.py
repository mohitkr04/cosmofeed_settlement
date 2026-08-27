"""
Comprehensive test suite for SEBI Registered Creator Discovery module.
Covers all 7 mandatory testing specifications from the project requirements:
  - Test 1: Normal Creator (no SEBI)
  - Test 2: SEBI Creator (badge/text match)
  - Test 3: VIG / Telegram Product detection
  - Test 4: Connected Creator recursive discovery
  - Test 5: Duplicate Creator deduplication & source merging
  - Test 6: API Failure error isolation
  - Test 7: Missing Data handling ('N/A', no fabricated values)
Plus Excel multi-sheet and CSV export verification.
"""

import os
import sys
import unittest
from unittest.mock import MagicMock, patch

# Ensure project root is on sys.path
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from sebi_creator_scraper.models import CreatorProfile, ProductEvidence
from sebi_creator_scraper.sebi_detector import detect_sebi_in_text, inspect_product_payload
from sebi_creator_scraper.queue_manager import QueueManager
from sebi_creator_scraper.id_loader import is_valid_creator_id, load_ids_from_reports
from sebi_creator_scraper.exporter import export_to_excel, export_to_csv
import openpyxl


class TestSEBIDiscoverySuite(unittest.TestCase):

    def setUp(self):
        self.mock_client = MagicMock()

    # -------------------------------------------------------------
    # Test 1: Normal Creator
    # -------------------------------------------------------------
    def test_01_normal_creator_no_sebi(self):
        """Creator with regular products and no SEBI registration -> SEBI Registered = NO."""
        cid = "66ea506c945e660013f90888"
        # Mock Kundli
        self.mock_client.get_creator_kundli.side_effect = lambda c, requested_action=None: {
            "status": True,
            "data": {
                "data": {
                    "_id": cid,
                    "Username": "regular_creator",
                    "Email": "reg@example.com",
                    "onboardedBy": "Team Alpha",
                    "vertical": "Education"
                }
            } if requested_action is None else {
                "data": {
                    "lastHundredSoldProducts": [
                        {
                            "_id": "6a0477ff3a48f8001346674b",
                            "productTtile": "Python Coding Bootcamp",
                            "productLink": "https://superprofile.bio/vp/6a0477ff3a48f8001346674b",
                            "productType": "page"
                        }
                    ]
                }
            }
        }
        # Mock Product Details
        self.mock_client.get_product_details.return_value = {
            "status": True,
            "data": {
                "pageData": {
                    "title": "Python Coding Bootcamp",
                    "description": "<p>Learn Python programming from scratch. No finance advice.</p>"
                }
            }
        }

        queue = QueueManager()
        profile = queue.process_one(cid, self.mock_client)

        self.assertEqual(profile.sebi_registered, "NO")
        self.assertEqual(profile.sebi_registration_number, "N/A")
        self.assertEqual(profile.status, "NOT_SEBI_REGISTERED")
        self.assertEqual(len(profile.product_evidence), 0)
        self.assertEqual(profile.products_checked, 1)

    # -------------------------------------------------------------
    # Test 2: SEBI Creator
    # -------------------------------------------------------------
    def test_02_sebi_creator_detection(self):
        """Creator whose product contains 'Registered with SEBI (INH000019099)' -> SEBI Registered = YES."""
        cid = "684158aa1c263c001338148a"
        self.mock_client.get_creator_kundli.side_effect = lambda c, requested_action=None: {
            "status": True,
            "data": {
                "data": {
                    "_id": cid,
                    "Username": "sebi_analyst",
                    "Email": "analyst@example.com",
                    "onboardedBy": "Baljeet Singh",
                    "vertical": "Finance"
                }
            } if requested_action is None else {
                "data": {
                    "lastHundredSoldProducts": [
                        {
                            "_id": "685f9401c345fa001398eb8d",
                            "productTtile": "Equity Research Advisory Service",
                            "productLink": "https://superprofile.bio/vp/685f9401c345fa001398eb8d",
                            "productType": "page"
                        }
                    ]
                }
            }
        }
        self.mock_client.get_product_details.return_value = {
            "status": True,
            "data": {
                "pageData": {
                    "title": "Equity Research Advisory Service",
                    "description": "<p>Daily stock recommendations. Registered with SEBI (INH000019099). All rights reserved.</p>"
                }
            }
        }

        queue = QueueManager()
        profile = queue.process_one(cid, self.mock_client)

        self.assertEqual(profile.sebi_registered, "YES")
        self.assertEqual(profile.sebi_registration_number, "INH000019099")
        self.assertEqual(profile.status, "SEBI_REGISTERED")
        self.assertGreaterEqual(len(profile.product_evidence), 1)
        self.assertIn("INH000019099", profile.sebi_evidence)

    # -------------------------------------------------------------
    # Test 3: VIG / Telegram Product Detection
    # -------------------------------------------------------------
    def test_03_vig_telegram_product_inspection(self):
        """Creator with VIG / Telegram product must NOT be excluded and SEBI must be detected."""
        cid = "67011223344556677889900a"
        self.mock_client.get_creator_kundli.side_effect = lambda c, requested_action=None: {
            "status": True,
            "data": {
                "data": {
                    "_id": cid,
                    "Username": "telegram_trader",
                    "Email": "trader@example.com",
                    "onboardedBy": "VIP Team",
                    "vertical": "Trading"
                }
            } if requested_action is None else {
                "data": {
                    "lastHundredSoldProducts": [
                        {
                            "_id": "67011223344556677889900b",
                            "productTtile": "Nifty Option Buying Telegram Channel",
                            "productLink": "https://superprofile.bio/vig/67011223344556677889900b",
                            "productType": "integratedGroup"
                        }
                    ]
                }
            }
        }
        # Product details response containing SEBI registration
        self.mock_client.get_product_details.return_value = {
            "status": True,
            "data": {
                "pageData": {
                    "title": "Nifty Option Buying Telegram Channel",
                    "productType": "vig",
                    "description": "<p>Exclusive Telegram Group. Registered with SEBI (INH000012345). Compliance approved.</p>"
                }
            }
        }

        queue = QueueManager()
        profile = queue.process_one(cid, self.mock_client)

        self.assertEqual(profile.sebi_registered, "YES")
        self.assertEqual(profile.sebi_registration_number, "INH000012345")
        self.assertEqual(profile.product_evidence[0].product_type, "vig")
        self.assertIn("vig", profile.product_evidence[0].product_link)

    # -------------------------------------------------------------
    # Test 4: Connected Creator Recursive Discovery
    # -------------------------------------------------------------
    def test_04_connected_creator_recursive_discovery(self):
        """Creator A connects to Creator B; Creator B has SEBI -> Both processed, Creator B identified as SEBI."""
        cid_a = "111111111111111111111111"
        cid_b = "222222222222222222222222"

        def mock_kundli(cid, requested_action=None):
            if cid == cid_a:
                if requested_action is None:
                    return {
                        "status": True,
                        "data": {
                            "data": {
                                "_id": cid_a,
                                "Username": "creator_a",
                                "connectedIds": [cid_b]
                            }
                        }
                    }
                return {"data": {"lastHundredSoldProducts": []}}
            elif cid == cid_b:
                if requested_action is None:
                    return {
                        "status": True,
                        "data": {
                            "data": {
                                "_id": cid_b,
                                "Username": "creator_b"
                            }
                        }
                    }
                return {
                    "data": {
                        "lastHundredSoldProducts": [
                            {
                                "_id": "333333333333333333333333",
                                "productTtile": "SEBI Premium Group",
                                "productLink": "https://superprofile.bio/vp/333333333333333333333333"
                            }
                        ]
                    }
                }
            return {"__error__": "not found"}

        self.mock_client.get_creator_kundli.side_effect = mock_kundli
        self.mock_client.get_product_details.return_value = {
            "status": True,
            "data": {
                "pageData": {
                    "title": "SEBI Premium Group",
                    "description": "<p>Registered with SEBI (INH000019099)</p>"
                }
            }
        }

        queue = QueueManager()
        queue.add_initial_creators({cid_a: {"Initial List"}})

        # Process Creator A
        prof_a = queue.process_one(queue.next_creator_id(), self.mock_client)
        self.assertEqual(prof_a.sebi_registered, "NO")
        self.assertIn(cid_b, prof_a.connected_creator_ids)

        # Creator B must have been enqueued
        self.assertTrue(queue.has_pending())
        next_id = queue.next_creator_id()
        self.assertEqual(next_id, cid_b)

        # Process Creator B
        prof_b = queue.process_one(next_id, self.mock_client)
        self.assertEqual(prof_b.sebi_registered, "YES")
        self.assertEqual(prof_b.sebi_registration_number, "INH000019099")
        self.assertIn(f"Connected Creator ID ({cid_a})", prof_b.discovery_sources)

    # -------------------------------------------------------------
    # Test 5: Duplicate Creator & Source Merging
    # -------------------------------------------------------------
    def test_05_duplicate_creator_merging(self):
        """Same Creator ID from multiple sources -> Unique record with merged discovery sources."""
        cid = "444444444444444444444444"
        queue = QueueManager()
        # Enqueue from multiple sources
        queue.add_initial_creators({cid: {"Audit (audit_2026-08-12.json)"}})
        queue.add_initial_creators({cid: {"Settlement Review"}})
        queue.enqueue_connected_id("999999999999999999999999", cid)

        self.assertEqual(len(queue.creators), 1)
        prof = queue.creators[cid]
        self.assertEqual(len(prof.discovery_sources), 3)
        self.assertIn("Audit (audit_2026-08-12.json)", prof.discovery_sources)
        self.assertIn("Settlement Review", prof.discovery_sources)
        self.assertIn("Connected Creator ID (999999999999999999999999)", prof.discovery_sources)

    # -------------------------------------------------------------
    # Test 6: API Failure Resilience
    # -------------------------------------------------------------
    def test_06_api_failure_resilience(self):
        """API failure on Creator A does not crash pipeline; Creator B processes successfully."""
        cid_fail = "555555555555555555555555"
        cid_ok = "666666666666666666666666"

        def mock_kundli(cid, requested_action=None):
            if cid == cid_fail:
                return {"__error__": "HTTP 500 Internal Server Error"}
            return {
                "status": True,
                "data": {
                    "data": {
                        "_id": cid_ok,
                        "Username": "ok_creator",
                        "Email": "ok@example.com"
                    }
                }
            }

        self.mock_client.get_creator_kundli.side_effect = mock_kundli
        self.mock_client.get_product_details.return_value = {"status": True, "data": {}}

        queue = QueueManager()
        queue.add_initial_creators({
            cid_fail: {"Source A"},
            cid_ok: {"Source B"}
        })

        prof_fail = queue.process_one(queue.next_creator_id(), self.mock_client)
        self.assertEqual(prof_fail.status, "ERROR")
        self.assertIn("500", prof_fail.error_message)

        prof_ok = queue.process_one(queue.next_creator_id(), self.mock_client)
        self.assertEqual(prof_ok.status, "NOT_SEBI_REGISTERED")
        self.assertEqual(prof_ok.username, "ok_creator")

        self.assertEqual(queue.stats["errors"], 1)
        self.assertEqual(queue.stats["processed"], 2)

    # -------------------------------------------------------------
    # Test 7: Missing Data Handling
    # -------------------------------------------------------------
    def test_07_missing_data_defaults_to_na(self):
        """Creator with missing email/vertical/onboardedBy sets 'N/A', never fabricates values."""
        cid = "777777777777777777777777"
        self.mock_client.get_creator_kundli.return_value = {
            "status": True,
            "data": {
                "data": {
                    "_id": cid,
                    "Username": "sparse_creator"
                    # Email, onboardedBy, vertical missing
                }
            }
        }
        self.mock_client.get_product_details.return_value = {"status": True, "data": {}}

        queue = QueueManager()
        profile = queue.process_one(cid, self.mock_client)

        self.assertEqual(profile.username, "sparse_creator")
        self.assertEqual(profile.email, "N/A")
        self.assertEqual(profile.onboarded_by, "N/A")
        self.assertEqual(profile.onboarding_vertical, "N/A")

    # -------------------------------------------------------------
    # Test 8: Multi-sheet Excel & CSV Export
    # -------------------------------------------------------------
    def test_08_excel_and_csv_export(self):
        """Validate Excel two-sheet export and CSV export."""
        tmp_xlsx = os.path.join(PROJECT_ROOT, "scratch", "test_out.xlsx")
        tmp_csv = os.path.join(PROJECT_ROOT, "scratch", "test_out.csv")

        c1 = CreatorProfile(
            creator_id="111111111111111111111111",
            username="analyst1",
            email="a1@example.com",
            onboarded_by="Manager X",
            onboarding_vertical="Finance",
            sebi_registered="YES",
            sebi_registration_number="INH000019099",
            sebi_evidence="Registered with SEBI (INH000019099)",
            discovery_sources={"Audit", "Settlement Review"},
            connected_creator_ids={"222222222222222222222222"},
            product_evidence=[
                ProductEvidence(
                    product_id="prod_101",
                    product_name="Pro Advisory",
                    product_type="page",
                    product_link="https://superprofile.bio/vp/prod_101",
                    sebi_registration_text="Registered with SEBI (INH000019099)",
                    sebi_registration_number="INH000019099",
                    evidence_source="Product Page Description",
                    has_sebi=True
                )
            ],
            status="SEBI_REGISTERED"
        )
        c2 = CreatorProfile(
            creator_id="222222222222222222222222",
            username="regular1",
            sebi_registered="NO",
            discovery_sources={"Connected ID"}
        )

        excel_res = export_to_excel([c1, c2], output_path=tmp_xlsx)
        csv_res = export_to_csv([c1, c2], output_path=tmp_csv)

        self.assertTrue(os.path.exists(tmp_xlsx))
        self.assertTrue(os.path.exists(tmp_csv))

        # Check Excel sheets
        wb = openpyxl.load_workbook(tmp_xlsx)
        self.assertIn("SEBI Registered Creators", wb.sheetnames)
        self.assertIn("SEBI Evidence - Products", wb.sheetnames)

        ws1 = wb["SEBI Registered Creators"]
        self.assertEqual(ws1.max_row, 3)  # Header + 2 creators

        ws2 = wb["SEBI Evidence - Products"]
        self.assertEqual(ws2.max_row, 2)  # Header + 1 product evidence row


if __name__ == "__main__":
    unittest.main()

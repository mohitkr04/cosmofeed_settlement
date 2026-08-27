"""
Unit tests for non_sebi_manager.py
==================================
Tests deduplication, cumulative volume calculations, and export generation.
"""

import os
import unittest
import tempfile
import json
import shutil

import non_sebi_manager as nsm


class TestNonSebiManager(unittest.TestCase):

    def setUp(self):
        # Create temp dir to isolate tests
        self.test_dir = tempfile.mkdtemp()
        self.orig_json = nsm.LEDGER_JSON_FILE
        self.orig_excel = nsm.CUMULATIVE_EXCEL_FILE
        self.orig_csv = nsm.CUMULATIVE_CSV_FILE
        self.orig_html = nsm.MANAGER_HTML_REPORT
        self.orig_pdf = nsm.MANAGER_PDF_REPORT

        nsm.LEDGER_JSON_FILE = os.path.join(self.test_dir, "test_ledger.json")
        nsm.CUMULATIVE_EXCEL_FILE = os.path.join(self.test_dir, "test_cumulative.xlsx")
        nsm.CUMULATIVE_CSV_FILE = os.path.join(self.test_dir, "test_cumulative.csv")
        nsm.MANAGER_HTML_REPORT = os.path.join(self.test_dir, "test_report.html")
        nsm.MANAGER_PDF_REPORT = os.path.join(self.test_dir, "test_report.pdf")

    def tearDown(self):
        nsm.LEDGER_JSON_FILE = self.orig_json
        nsm.CUMULATIVE_EXCEL_FILE = self.orig_excel
        nsm.CUMULATIVE_CSV_FILE = self.orig_csv
        nsm.MANAGER_HTML_REPORT = self.orig_html
        nsm.MANAGER_PDF_REPORT = self.orig_pdf
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def test_strict_deduplication_across_days(self):
        """Test that same creator across multiple dates accumulates without duplicating."""
        creator_id = "6457e6db7084bb0020b8212d"

        # Day 1: 27 Aug (Payout 5,000)
        day1_batch = [
            {
                "creatorId": creator_id,
                "username": "test_creator_1",
                "payoutAmount": 5000.0,
                "telegramIntegration": True,
                "telegramEligible": True,
                "sebiRegisteredNo": "No",
                "telegramProductId": "prod_day1"
            }
        ]
        nsm.record_daily_settlements(day1_batch, audit_date="2026-08-27")

        ledger1 = nsm.load_ledger()
        self.assertEqual(len(ledger1["creators"]), 1)
        c1 = ledger1["creators"][creator_id]
        self.assertEqual(c1["cumulativePayoutVolume"], 5000.0)
        self.assertEqual(c1["latestPayoutAmount"], 5000.0)
        self.assertEqual(c1["daysFlaggedCount"], 1)
        self.assertEqual(c1["datesObserved"], ["2026-08-27"])

        # Day 2: 28 Aug (Payout 3,500 for same creator, plus 1 new creator)
        day2_batch = [
            {
                "creatorId": creator_id,
                "username": "test_creator_1",
                "payoutAmount": 3500.0,
                "telegramIntegration": True,
                "telegramEligible": True,
                "sebiRegisteredNo": "No",
                "telegramProductId": "prod_day2"
            },
            {
                "creatorId": "63bea67b4bb82e0033aee831",
                "username": "test_creator_2",
                "payoutAmount": 2000.0,
                "telegramIntegration": True,
                "telegramEligible": True,
                "sebiRegisteredNo": "No",
                "telegramProductId": "prod_2"
            }
        ]
        nsm.record_daily_settlements(day2_batch, audit_date="2026-08-28")

        ledger2 = nsm.load_ledger()
        # MUST have exactly 2 unique creators, NOT 3
        self.assertEqual(len(ledger2["creators"]), 2, "Expected strictly 2 unique creators without duplicates")

        c1_updated = ledger2["creators"][creator_id]
        # Payout should be cumulative 5,000 + 3,500 = 8,500
        self.assertEqual(c1_updated["cumulativePayoutVolume"], 8500.0)
        self.assertEqual(c1_updated["latestPayoutAmount"], 3500.0)
        self.assertEqual(c1_updated["daysFlaggedCount"], 2)
        self.assertEqual(c1_updated["datesObserved"], ["2026-08-27", "2026-08-28"])
        self.assertIn("prod_day1", c1_updated["telegramProductIds"])
        self.assertIn("prod_day2", c1_updated["telegramProductIds"])

    def test_file_exports_generated(self):
        """Test that Excel, CSV, HTML, and PDF are generated."""
        batch = [
            {
                "creatorId": "6457e6db7084bb0020b8212d",
                "username": "creator_sample",
                "payoutAmount": 1500.0,
                "telegramIntegration": True,
                "telegramEligible": True,
                "sebiRegisteredNo": "No",
                "telegramProductId": "p123"
            }
        ]
        nsm.record_daily_settlements(batch, audit_date="2026-08-27")

        self.assertTrue(os.path.exists(nsm.CUMULATIVE_EXCEL_FILE))
        self.assertTrue(os.path.exists(nsm.CUMULATIVE_CSV_FILE))
        self.assertTrue(os.path.exists(nsm.MANAGER_HTML_REPORT))
        self.assertTrue(os.path.exists(nsm.MANAGER_PDF_REPORT))

        # Check Excel content
        import openpyxl
        wb = openpyxl.load_workbook(nsm.CUMULATIVE_EXCEL_FILE)
        self.assertIn("Executive Summary", wb.sheetnames)
        self.assertIn("Non-SEBI Master Registry", wb.sheetnames)


if __name__ == "__main__":
    unittest.main()

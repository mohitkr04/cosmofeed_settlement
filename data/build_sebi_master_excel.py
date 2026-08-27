"""
Compiles the verified SEBI-registered creators from the organizational PDF into sebi_master_creators.xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(HERE)

# Complete raw data transcribed from the 22-page PDF
MASTER_ENTRIES = [
    # Page 1
    ("67bff2a289232b0013343cec", "RAMPRASAD OMPRAKASH MUNDADA", "INH000010690", "maxprocapitaladvisory@gmail.com", "TGI"),
    ("61e996f678f4fa74e5da637d", "61e996f678f4fa74e5da637d", "INH000019099", "sadityagatep@gmail.com", "TGI"),
    ("679df13237b926001364797c", "stockgainer", "INH100007879", "kapilverma03@gmail.com", "TGI"),
    ("628c71b19a6375439f664285", "P Investments", "INH000017471", "Piyushheda.ph@gmail.com", "TGI"),
    ("6298da07fa0bd672300dbe1f", "Stockcepe", "INH000016481", "gunjan00281@gmail.com", "TGI"),
    ("67066dc98bfbb400131c09e1", "Equiscan Research", "INH000013183", "equiscan24@gmail.com", "TGI"),
    ("6397007d7f42a30034612430", "Alok Daiya", "INH000011468", "raalokdaiya@gmail.com", "TGI"),
    ("65f060d70456340019a780ad", "Krishna Kumar Malu", "INH000010380", "Krishna.malu.ksm@gmail.com", "TGI"),
    ("63078229747211211991aeae", "Vijay Lakshmi", "INH000009782", "stocksmarkettoday@gmail.com", "TGI"),
    ("6745ab3c36818b0013b2ab20", "Vivek KR Singh", "INH000010812", "viveksinghlko24@gmail.com", "TGI"),
    # Page 2
    ("679cafc8e903520013caf8ef", "cmagurvinder", "INH000017480", "gurvinder.malhotra82@gmail.com", "TGI"),
    ("6478f24bd858f1002029d016", "The Trading Circle", "INH000019017", "kaushikghosh014@gmail.com", "Unspecified"),
    ("63dd2203d410d20034e748c2", "S.K WEALTH ADVISORY", "INH000019187", "Deepak.khatri9585@gmail.com", "Unspecified"),
    ("67c59ae71803020013622b1d", "Aditya Umesh Hujband", "INH000011185", "mbainvestmentwala@gmail.com", "PP"),
    ("67d02ac3369e880013bf46c6", "SFP Research", "INH000018124", "sfpresearch24@gmail.com", "TGI"),
    ("67b8b0e3b45eb001367323b", "trade milan", "INH000019327", "spkumar.researchanalyst@gmail.com", "TGI"),
    ("62a87a8907c75110e6ec2a70", "Kedia Research", "INH000009667", "contact@kediaresearch.com", "PP"),
    ("67060dac6a83ed0013d5c14c", "Vaibhav Pandey", "INH000017347", "vaibhavpandey16188@gmail.com", "TGI"),
    ("63456b020f75ab3c048d046d", "optiontraderadda", "INH000019026", "ra.rohansahu@gmail.com", "TGI"),
    ("67eb8f3ebca74f00139aafa7", "Trudence Capital Advisors Private Limited", "INH000015747", "compliance-grievance@trudencecapital.com", "PP"),
    ("67e78681f903980013aa43e7", "Tanya", "INH000020147", "stockmatrixresearch@gmail.com", "TGI"),
    # Page 3
    ("6735d02d1e2c9e00139b37ef", "Telgue Trader shyam", "INH000018869", "candwicktrader@gmail.com", "Unspecified"),
    ("674c34f3b705d90013e101a5", "wahab", "INH000016029", "Wahab.analyst@gmail.com", "Unspecified"),
    ("620bacd7958992296a26d7e8", "Anshul Aggrawal", "INH000016092", "ANSUL147@GMAIL.COM", "TGI"),
    ("675fdb7b5794a00013679bcd", "Pawan", "INH000019053", "greencandle.research@gmail.com", "Unspecified"),
    ("67a9ad832b922d0013738054", "Equiscan Research", "INH000013183", "equiscan24@gmail.com", "Unspecified"),
    ("6389a42f9e81360031bca085", "Ram Mundada", "INH000010690", "Mundada79@gmail.com", "Unspecified"),
    ("6404527175b2e000200a8c5c", "Sidharth Rajput", "INH000016612", "Info@thestocksavvy.com", "TGI"),
    ("659fe1fee66536001f0bf11c", "Sheetal Kunder", "INH000013800", "skunder2727@gmail.com", "TGI"),
    ("65db4cd5c61ce00013dfb42e", "sebiravineetsaxena", "INH000013855", "vineetsaxena09@gmail.com", "Unspecified"),
    ("67f4ad39184d330013117481", "Priti Tiwari", "INH000019822", "priti.tiwari01@gmail.com", "PP"),
    ("67ef99b9fd01df0013ae9500", "commodityquant SEBI RA", "INH200005212", "mthara03@gmail.com", "PP"),
    ("62e0048df83d933bcd624f83", "Systematic Traders", "INH000020156", "cosmofeedvivek@gmail.com", "TGI"),
    # Page 4
    ("67fc90f8cd18010013e953e6", "Pramod Patil", "INH000017286", "pramodpatil3385@gmail.com", "TGI"),
    ("6801d744c5314a0013fc527b", "logesh kumar", "INH200004648", "masterstocktips@gmail.com", "TGI"),
    ("67ebbb3e331c1f001407c1fc", "shyam", "INH000018735", "support@srequisearch.com", "TGI"),
    ("619616196b48113c833f9733", "HITESHWARI KUMARI", "INH000016472", "INFO@BREAKOUTMANTRA.COM", "TGI"),
    ("680737d29a59d00013ed333b", "Amit Bhattacharjee", "INH000019886", "aalgobreaths@gmail.com", "TGI"),
    ("68061811c3716b00137f76c7", "stock ocean", "INH000011006", "vardhanharsh548@gmail.com", "PP"),
    ("65aacfa14b8a0b001e456f6d", "Rahul chandra", "INH000020059", "mylittles88371@gmail.com", "TGI"),
    ("6808bad7c0556200138e5d85", "Ram Mundada", "INH000010690", "capitaladvisorymaxpro@gmail.com", "TGI"),
    ("67e28569a9248e001318b29d", "Vaibhav Sandipan Mote", "INH000019895", "motefinn@gmail.com", "TGI"),
    ("679e34f9c7375300134f5a45", "Ankit gupta", "INH100007231", "wealthconsultant.ra@gmail.com", "PP"),
    ("649017bdf584c600204d1c33", "wisebull", "INH100009992", "Wisebull20@gmail.com", "TGI"),
    # Page 5
    ("61d2ef7847e88d778cef2c4b", "Sudhir", "INH300008474", "info@divinecapitalmarket.com", "TGI"),
    ("683176582fdccd00132d74d4", "Arth Sadhana", "INH000020341", "arthsadhana.rnc@gmail.com", "TGI"),
    ("6832c6b611fc5b001340a475", "Pradip halder", "INH000016126", "researchdesk@phdcapital.in", "TGI"),
    ("626cbc7e460460659d1eae0f", "Sankalp", "INH000014119", "vvsankalp@gmail.com", "TGI"),
    ("6352bafe2a0a310cbea88f69", "CA Akash Garg", "INH000011501", "caakashgarg97@gmail.com", "PP"),
    ("66db5f8013d5510013fd0388", "vyomresearch", "INH000018221", "Compliance@vyomresearch.in", "TGI"),
    ("667e674fcdab3f00144d2e8c", "Prameela", "INH000016074", "info@gainxcapital.in", "TGI"),
    ("6810c7bc5f7fa000135017b8", "dhaval", "INH000002855", "dhavalpvyas@dhavalpvyas.com", "PP"),
    ("6856780ac5a2ba00132ff991", "Mohd Rizwan", "INH000020536", "rizwan.cia@gmail.com", "TGI"),
    ("66433629ed4ca000136c0f2f", "Mohit Agarwal", "INH000011954", "intradaymatch008@gmail.com", "TGI"),
    ("6383c737e54258003ebcaf6c", "MAUSAM NAGPAL", "INH000020554", "support@nagpalwealthadvisor.com", "TGI"),
    # Page 6
    ("67f8e11b996d5b0013b31dc8", "Nandita Rai", "INH000018054", "support@bhaktiwealth.com", "PP"),
    ("685552989c2cef0013354541", "Ruchita Shah", "INH000016995", "rsresearchanalyst@gmail.com", "TGI"),
    ("64718eae764e8500218d75d5", "Saras Jain", "INH000013086", "sarasbadrikedar@gmail.com", "TGI"),
    ("68594aad83c8bb0013947a89", "Aakanksha Gupta", "INH000013457", "aakanshauptti@gmail.com", "TGI"),
    ("670e65c64bceb70013de1d42", "Swapnil shinde", "INH000018504", "optionoraclebusiness@gmail.com", "TGI"),
    ("68678c9f552560001215d4ee", "Divine Capital", "INH000021243", "contact@equivestresearch.com", "TGI"),
    ("686cc61e09eb2c0013194347", "bULLS sTRATEGY", "INH000019558", "mart53962@gmail.com", "TGI"),
    ("6234c14002100543e97a0ac5", "Chirag jain", "INH000014298", "jain.chirag3738@gmail.com", "TGI"),
    ("64e35f864a24de002b81a873", "nitesh jain", "INH000021614", "jain.nitesh0702@gmail.com", "TGI"),
    ("643d1d81cb7c2e0020fcdf50", "Naveen Kumar Choudhary", "INH000011088", "eftiwealth@gmail.com", "TGI"),
    ("6384b3611d2957002c6332a5", "RAJ PARMAR", "INH000021535", "rajparmar1120@gmail.com", "TGI"),
    ("66712522a0ac7a001342510c", "Latesh Narula", "INH000016269", "shriramtradersclub01@gmail.com", "TGI"),
    # Page 7
    ("68822b829172ab0013e42f43", "Bivhav Nayak", "INH000010849", "bibhav.nayak@gmail.com", "TGI"),
    ("655e33765187d2001e865423", "Vishnu Deekonda", "INH200009500", "investordiary.vishnu@gmail.com", "TGI"),
    ("66289aacc0d66c0013edeb77", "CHART COLONY ENTERPRISES", "INH000018018", "compliancechartcolony@gmail.com", "TGI"),
    ("67c7d9335fe3f40013cc887d", "Kulneet Singh Bindra", "INH000014845", "traderpaaji@gmail.com", "TGI"),
    ("67f3c7bd42528b0013ca9dca", "ABISHEK KONDAGUNTA VENKATESH", "INH000022066", "abishekkvenkatesh@outlook.com", "PP"),
    ("688d6aa348cd3e0013d3c177", "Shashank udupa", "INH000021207", "shanks.udupa@gmail.com", "TGI"),
    ("688c773f69c0b30013edad3b", "Ashutosh Kabra", "INH000015349", "ashurj.maheshwari@gmail.com", "TGI"),
    ("627d4bbd91125d5ba8a436cf", "RAJESHWAR MANDAL", "INH000018878", "rajeshwar.mandal97@gmail.com", "TGI"),
    ("68948280c886ee001429818c", "Sourabh Bajaj", "INH000013378", "sourabhbajaj01@gmail.com", "TGI"),
    ("623c88f855611802acdcb6e0", "Sourabh Bajaj", "INH000013378", "cabajaj.sma@gmail.com", "TGI"),
    # Page 8
    ("6895d18d605f210013808fde", "Vibhor Gupta", "INH000019257", "vibhoregupta@gmail.com", "TGI"),
    ("68947929ec57170013bfec2e", "SHUBHAM", "INH000020411", "goyal2301shubham@gmail.com", "TGI"),
    ("68945c54ec57170013bde864", "Vaibhav", "INH000018513", "Vaibhavmiyani9@mail.com", "TGI"),
    ("67543a6e052d2500138f3dba", "vishal", "INH000016816", "vishaltrehan_ra_inh000016816@outlook.com", "TGI"),
    ("65df0d8863d72a0013f718d1", "Madhu Bansal", "INH000010672", "madhubansal0987@gmail.com", "TGI"),
    ("689db837eb2a2d0013830aeb", "Shikha Kapur", "INH000019169", "shikhaa.kapur@gmail.com", "PP"),
    ("68a01ec6bfa433001304d708", "Bidyut Biswas", "INH000014368", "learntinvest@gmail.com", "TGI"),
    ("68a42823e5e53800130c2578", "SEIRI INVESTMENT PRIVATE LIMITED", "INH000022765", "nitendrajhariya999@gmail.com", "TGI"),
    ("654bbf0546a134001e3bc1e4", "Kirti Saraf", "INH000014784", "kirti.saraf11@gmail.com", "TGI"),
    ("68a843962739f900133d79ee", "LALIT KUMAR MUNDHRA", "INH000017091", "sarthisfinancial@gmail.com", "TGI"),
    # Page 9
    ("6655e6c31474ac00130b8d2c", "stock sovet", "INH000016366", "support@stocksovet.com", "TGI"),
    ("689afb1c534cb2001356ede2", "Kiran Bala", "INH000016445", "learntradingwithkb@gmail.com", "TGI"),
    ("66865c4bd0580400135bab29", "Samir Kumar", "INH000022914", "singhsamir2611@gmail.com", "TGI"),
    ("666d5ede6667a30013276f7b", "Rohit", "INH000022543", "ra-support@retirewithrohit.com", "Unspecified"),
    ("67d687df0e55b30013c2dada", "Stockbox Technologies private limited", "INH100008799", "stockboxtech@gmail.com", "Unspecified"),
    ("68b3ec1d54d5b600132183eb", "ALDERLEAF STOCKMANTRA PRIVATE LIMITED", "INH000019099", "sadityagatep@gmail.com", "PP"),
    ("68abc9ff6ecb490013e98579", "DG Share Market Research pvt. ltd.", "INH000015534", "dgsharemarketresearch@gmail.com", "PP"),
    ("68b52dbd62e463001387fdc2", "DG Share Market Research pvt. ltd.", "INH000015534", "dgsharemarketresearch@gmail.com", "PP"),
    ("6373ded643e1100032688133", "Umesh Sharma", "INH000015622", "ankitgupta2669@gmail.com", "TGI"),
    ("68b43d395794160013355493", "ASHISH KUMAR", "INH000011866", "gargashish585@gmail.com", "TGI"),
    # Page 10
    ("68a95844634b62001343c9a1", "Anand Sharadchandra Pathak", "INH000020970", "anandpathakra@gmail.com", "TGI"),
    ("67163375f900130013ef2fee", "Kavita", "INH000017134", "kavitacontact2024@gmail.com", "TGI"),
    ("65adfbc13815fa001ec48bcf", "Sumit Bapusaheb Kamblaykar", "INH000023047", "Rasaumit@gmail.com", "PP"),
    ("68b549d580475700138bbb9f", "Rajnikant Vinubhai Chauhan", "INH000021988", "swaminarayantradingcompany@gmail.com", "TGI"),
    ("68c503224351760013548b52", "Krishna Murti", "INH000016977", "kmresearchanalyst@gmail.com", "TGI"),
    ("68c7a05ab1b06c00134ea4ac", "JAY B HAWALDAR", "INH000016685", "hawaldarjay@gmail.com", "TGI"),
    ("663bacb0f458af0013701c58", "Mahyavanshi Jignesh Jayantibhai", "INH000022376", "contact@stoxsheekho.com", "TGI"),
    ("6400835edff56800208036e8", "PANKAJ KUMAR JAIN", "INH000021386", "support@eyeontrade.com", "TGI"),
    ("68c52bdac8ee3a0013d01089", "VIJAY KUMAR GUPTA", "INH000020226", "RESEARCHANALYSTVIJAY@GMAIL.COM", "TGI"),
    ("68c912d18dc9ea0013ec9a10", "Priyam Mehta", "INH000019239", "Gapupequity@gmail.com", "TGI"),
    ("68c114b83fdf1300136caa7d", "GOPAL LOHAR", "INH000011963", "gopal25lohar@gmail.com", "TGI"),
    # Page 11
    ("659c38b5d3c938001e685abe", "Sourabh Bajaj", "INH000013378", "cabajaj.sma@gmail.com", "TGI"),
    ("68d0ca09287cca001359dd93", "Vivek Kumar", "INH000021915", "Vivekmalik3003@icloud.com", "TGI"),
    ("68c53cb6c8ee3a0013d207f2", "Binita Bharti", "INH000021456", "binitab21202gmail.com", "TGI"),
    ("64fb0ee375d3bf001ea52867", "Priyam Mehta", "INH000014155", "Contact@ascendwealth.in", "TGI"),
    ("666abae11022650014e828f0", "Sunil Gurjar", "INH000014261", "sunilgurjar7971@gmail.com", "TGI"),
    ("666e8a3cf52ecb0013f648c2", "Pravin Khetan", "INH000019521", "pravinkhetan@gmail.com", "TGI"),
    ("68d28808ad1b82001393be28", "Afaque Ali", "INH000015011", "afaque4yu@gmail.com", "TGI"),
    ("6399af15f314c9002c78ccee", "Sami Ahmad Sajjad", "INH000020819", "samisajjad786@gmail.com", "TGI"),
    ("6251417a110f1d34415645ff", "Hitesh Somani", "INH000010645", "hiteshsomani84@gmail.com", "TGI"),
    ("68da3d9b97a50f0013bbe8df", "Thacker Bhavesh H", "INH000020907", "dr.bhaveshthacker@gmail.com", "TGI"),
    ("630b39efba45c4661500a699", "Birendra Kumar Pandey", "INH000020217", "armando.bhw@gmail.com", "TGI"),
    ("68df70476e897500136b5591", "BIBEKANANDA ROY", "INH000023250", "royresearchanalyst@gmail.com", "TGI"),
    # Page 12
    ("63ece050bd702e002bbaa257", "SANYAM CHOPRA", "INH000014137", "mfs.sebi.ra@gmail.com", "PP"),
    ("67f391d32bd44200137da3b9", "JAI SHEWARAMANI", "INH000010681", "jai.shewaramani1@gmail.com", "TGI"),
    ("68c3e79f3812dd00134aaf25", "LOGIC TRADE ENTERPRISES", "INH000023524", "logictrade.co.in@gmail.com", "TGI"),
    ("631f3805a293b361d9224396", "Birendra Kumar Pandey", "INH000020217", "vpcapitalresearchpvtltd@gmail.com", "TGI"),
    ("68e3ab99dc3c5a0013cf4b44", "DG Share Market Research pvt. ltd.", "INH000015534", "dgsharemarketresearch@gmail.com", "PP"),
    ("68e6056978871e00132bfd29", "DG Share Market Research pvt. ltd.", "INH000015534", "dgsharemarketresearch@gmail.com", "PP"),
    ("68e60b52cddfee00134a7888", "DG Share Market Research pvt. ltd.", "INH000015534", "dgsharemarketresearch@gmail.com", "PP"),
    ("68e60fe65b84d900139e1236", "DG Share Market Research pvt. ltd.", "INH000015534", "dgsharemarketresearch@gmail.com", "PP"),
    ("68e613c6c61a7600130dc73c", "DG Share Market Research pvt. ltd.", "INH000015534", "dgsharemarketresearch@gmail.com", "PP"),
    # Page 13
    ("649d2c3900f8850020e8e2ca", "Kaushal Prafulla Somani", "INH000016418", "kpsomani15@gmail.com", "TGI"),
    ("64aeb067de6c780020f7d740", "Arthasiddhi Equity Advisory Services", "INH000021012", "arthasiddhiadvisory@gmail.com", "TGI"),
    ("684158aa1c263c001338148a", "Vibhor Varshney", "INH100005419", "nakulvibhor@gmail.com", "PP"),
    ("69036bf7ff27890013ff48bd", "Rachit Jain", "INH000022738", "ra.rachitjain@gmail.com", "TGI"),
    ("690384c4727f9f0013a01d03", "pradeep carpenter", "INH000019309", "rapradeepcarpenter@gmail.com", "TGI"),
    ("64c22361ec0597001f1a95c6", "Umang Bhutada", "INH000023001", "umangbhutada100@gmail.com", "TGI"),
    ("68e0cf03e32be10013fc67e0", "Vinit Aggarwal", "INH000016676", "vinitk.aggarwal@gmail.com", "TGI"),
    ("6902fcb2e9e9a30013509224", "Vedant Kelkar", "INH000019266", "princecapital09@gmail.com", "TGI"),
    ("690eeaabed740f00134ca76e", "Manish N Chetwani", "INH000023472", "manishchetwani@yahoo.com", "TGI"),
    ("67e68de7ecd8250013e58174", "Nishit Bipin Doshi", "INH000017198", "nishit189@gmail.com", "PP"),
    ("62c11a54f3090551e28230e6", "SAHIL VERMA", "INH000023463", "sahilverma70914@gmail.com", "PP"),
    ("66dbe68817809c0013fd1a4a", "TRADING BRAIN 4U", "INH000009418", "ansarinaved86@gmail.com", "PP"),
    # Page 14
    ("68e8986b8bda6500131a7b1c", "Aashish Bansal", "INH000023533", "caashishbansal17@gmail.com", "TGI"),
    ("68ef6096eb549e00137fa537", "M/s. Nexttech Securities Private Limited", "INH000021067", "nexttechsecurities@gmail.com", "TGI"),
    ("6927dc12b150fc00133999a8", "RISHI RAI", "INH000010423", "rishiraisaxena77@gmail.com", "TGI"),
    ("6894b5cded02fe0013b89f54", "M/S MAM AND KAV SECURITIES", "INH000022561", "mamandkavsecurities@gmail.com", "TGI"),
    ("6788721993df9b00130b9291", "Chahat Aggrawal", "INH000018577", "Chahatmangla.ra@gmail.com", "PP"),
    ("67066dc98bfbb400131c09e1", "Equiscan Research", "INH000013183", "equiscan24@gmail.com", "TGI"),
    ("693944560e277600137fc7c6", "Sanjoy Ghosh", "INH000021696", "sghoshmar2016@gmail.com", "Both"),
    ("692e7b3dd42b600013bd5a25", "shubham manojkumar jain", "INH000023782", "shuubhaamjain@gmail.com", "TGI"),
    ("693bab144fa2640012ae28b8", "Vedant Kelkar", "INH000019266", "shejwaljalindar155@gmail.com", "TGI"),
    ("6942a355a51d220012b47229", "Chaitanya Agarwal", "INH0000224641", "chaitanyaagarwal@gmail.com", "TGI"),
    ("6891e2d6e95ce00013d08761", "Purvesh Mehta", "INH000014748", "purvesh22@gmail.com", "TGI"),
    # Page 15
    ("64213282171195001fb5227c", "VYOM", "INH000018221", "Compliance@Vyomresearch.in", "TGI"),
    ("68cbd07c5e35850013de5264", "VYOM", "INH000018221", "Compliance@Vyomresearch.in", "TGI"),
    ("66db5f8013d5510013fd0388", "VYOM", "INH000018221", "Compliance@Vyomresearch.in", "TGI"),
    ("6585b49a55e4ed001e8b7cd3", "MANJEET KAUSHIK", "INH000017277", "mail.wolfclub@gmail.com", "TGI"),
    ("694f7b2833ceb800139600ee", "DEEPAK DHANWANI", "INH000023348", "masterchartists@gmail.com", "TGI"),
    ("668d2403bd8d0c001342122d", "APURVA ANNAVI", "INH000024064", "sebi.ra.apurva@gmail.com", "TGI"),
    ("6954db9e53660300131cc9bc", "Pawandeep", "INH000023904", "pm17pawandeps@iimidr.ac.in", "TGI"),
    ("67543af6d7e18b0013b20364", "ROHAN RAJENDRA AHER", "INH000023773", "rohan.aher08@gmail.com", "TGI"),
    ("656827afc86416001d935758", "Vedant Kelkar", "INH000019266", "wealthmantraaa@gmail.com", "TGI"),
    ("68a85b6eabd7d30013dc5a8c", "Shrikant Pandey", "INH000023986", "shrikantpandey781@gmail.com", "TGI"),
    ("66500de8ce5811001334de33", "Neha Dogra", "INH000010195", "nehasharma66689@gmail.com", "PP"),
    # Page 16
    ("696f1a1b1810ca001439ffd8", "VIJAYA KUMAR SUBRAMANIYAN", "INH000013688", "capitalriseresearchservices@gmail.com", "PP"),
    ("69705c4321d251001320f200", "Gourav Jain", "INH000024268", "Investorsempire1111@gmail.com", "PP"),
    ("667cde6d8fa7f300136d8f1c", "Dinesh Kumar", "INH000024152", "dinesh.research.ra@gmail.com", "TGI"),
    ("69672e197065cd00133ddad0", "AKHIL KUMAR RAI PROPRIETOR 3i RESEARCH", "INH000012689", "akhilkumar.r@outlook.com", "TGI"),
    ("697dbf0da47c6800135b1b9f", "ASHISH SINGH", "INH000014340", "ashish.singh@geniresearch.com", "TGI"),
    ("64ede6651dba71001ef6e381", "GONDALIYA BHAVESH CHHAGANBHAI", "INH000024772", "bhaveshgondaliya1993@gmail.com", "TGI"),
    ("6982eac2fae9950013901bfb", "Pramod Patil", "INH000017286", "pramodpatil3385@gmail.com", "TGI"),
    ("692eb6d120bfc5001378d10a", "Vivek Lochan Sharma", "INH000023083", "viveklochansharma@gmail.com", "TGI"),
    ("691f59a0f310eb0013272119", "Darshan N Rao", "INH000017958", "raodarshan@gmail.com", "TGI"),
    # Page 17
    ("6989a6bb75efa100134e59d8", "DIPAK M TAKODARA", "INH000013679", "dipak.takodara@gmail.com", "TGI"),
    ("67060dac6a83ed0013d5c14c", "vaibhav pandey", "INH000017347", "vaibhavpandey16188@gmail.com", "TGI"),
    ("698ec8ff316e00001382adca", "Rounaq Bakshi", "INH000022747", "Bakshirounaq5@gmail.com", "PP"),
    ("62b30fec6985f90a6ea83d0c", "BLUEMOON RESEARCH & FINANCIAL SERVICES", "INH000016135", "Ma94noj@gmail.com", "TGI"),
    ("699847f87584650013a7b4df", "Ritesh Gupta", "INH000014775", "guptaritesh120787@gmail.com", "PP"),
    ("69633804fc5f2b0013dc46a9", "Ritesh Gupta", "INH000014775", "riteshgupta3202@gmail.com", "PP"),
    ("6867a8c7af0a3b0013826001", "Nemesh Pandey", "INH000018674", "nemeshpandey@gmail.com", "PP"),
    ("699fd3dd4de65100136e8c5f", "Na Supreme Research Private Limited", "INH000017745", "nsen1525@gmail.com", "PP"),
    ("67eaed8647e20c0013086e26", "Chintan Bajaj", "INH000023649", "Connect.nirvanaresearch@gmail.com", "TGI"),
    ("69b50f0dafac6d001348bde0", "YASH CHOUHAN", "INH000025355", "yashchauhanresearchanalyst@gmail.com", "TGI"),
    ("69b7f2ebfea3070014ac8bb0", "Manas Budhirja", "INH000025027", "budhirajamanas23@gmail.com", "PP"),
    # Page 18
    ("6990677a1a62e50013864607", "SIDDHARTH BHANUSHALI ADVISORY PRIVATE LIMITED", "INH000022905", "krantis@sidsnb.com", "PP"),
    ("67f8e11b996d5b0013b31dc8", "Nandita Rai", "INH000018054", "support@bhaktiwealth.com", "TGI"),
    ("69c82c3186513100135df724", "NV TARUN GOUTHAM", "INH000025443", "tarun.goutham@gmail.com", "TGI"),
    ("6923e79a38a1c20013513b2d", "Om Nitin Patil", "INH000024587", "omnitin2000@gmail.com", "TGI"),
    ("66fbc1c29002d40012d868d8", "Aishani Dhawan", "INH000020703", "aishanidhawanofficial@gmail.com", "TGI"),
    ("69d739c50b84e10013990515", "Rishi Wadhawan", "INH000025726", "wadhawan_sonu@rediffmail.com", "TGI"),
    ("667c0a50a635ea0014b1389c", "NITESHKUMAR SIDDHRAM SHENDRE", "INH000015020", "nsendre@gmail.com", "TGI"),
    ("68b291142afe1f0013f35d90", "TAURUS ANALYTICS", "INH000025276", "yeolekarrishabh@gmail.com", "TGI"),
    ("69e1cd5f2d292d00136f412f", "ALDERLEAF STOCKMANTRA PRIVATE LIMITED", "INH000019099", "alderleafstockmantrapvtltd@gmail.com", "TGI"),
    ("69e73e04d2582e0013922cb0", "Ravinder Sharma", "INH000024578", "trade.verse2407@gmail.com", "TGI"),
    # Page 19
    ("678bb46834ce7300134587d4", "Vivek Singh Rajpoot", "INH000024550", "galaxyvivek4@gmail.com", "TGI"),
    ("6976e15a59943300133c40df", "VARUN JOSHI", "INH000026141", "support@nirivantes.com", "TGI"),
    ("69c22cdd76f39700134825dd", "Missing Name", "INH000024921", "stoxplexresearch@gmail.com", "TGI"),
    ("6969234b76dc8f001352a9b8", "SAURABH MITTAL", "INH000022835", "Ra.saurabhmittal@gmail.com", "TGI"),
    ("69f5991ba8c2c6001350bf88", "Tapas Chandra Baskey", "INH000026381", "tapaschandraofficial@gmail.com", "TGI"),
    ("69b911031bb9f60013874054", "Vishal Jagannath Khapre proprietor of SafeInvex Research", "INH000025036", "vishu.khapre@gmail.com", "TGI"),
    ("6914066cd56f53001329d27e", "Suhani Sandip Katariya", "INH000026257", "casuhanikatariya@gmail.com", "TGI"),
    ("6a0407939c43f10013e1e812", "H K RIYAZ SALAM PROPRIETOR OF EQUITY ALFA ANALYTICS", "INH000024985", "riyazsalam.hk@mail.com", "Both"),
    ("6a0ac87918bbd70013b11c18", "ARUN S VIJAY", "INH000020633", "arun@alphaturtles.in", "TGI"),
    ("6a118831557fe60013349dbd", "FINLIGHT RESEARCH INDIA PRIVATE LIMITED", "INH000013518", "ra.finlightresearch@gmail.com", "TGI"),
    # Page 20
    ("6a102af970f880001322b2c4", "SANDESH KEDIA", "INH000020882", "kedia.sandesh@gmail.com", "TGI"),
    ("6a0850d437a61300131fe750", "Sourav Kushwah", "INH000025823", "aoneresearch1@gmail.com", "TGI"),
    ("6a0c427d1be92d0013bdfdab", "6a0c427d1be92d0013bdfdab", "INH000020697", "swativijay021@gmail.com", "TGI"),
    ("62eb6ce9ea2f933920d5cf73", "Bullish India Capital Pvt. Ltd.", "INH000010070", "info@bullishindia.com", "TGI"),
    ("6a107b341c48660013c81d1e", "JITENDRA SINGH", "INH000027335", "jitendraresearchanalyst@gmail.com", "TGI"),
    ("66af4c4b22d2650013570023", "Shivendru shekhar", "INH000027469", "shivendru39@gmail.com", "TGI"),
    ("6a27e1de91182c0013b02261", "RAVI BHARATKUMAR GUPTA", "INH000012184", "opopoceansofprofit@gmail.com", "TGI"),
    ("6536a25a74d76b001d774459", "Omkar Devram Bhalsing", "INH000025142", "Omkarbhalsing@gmail.com", "PP"),
    ("689338363085aa0013cadca4", "Khushal Growwealth Research Private Limited", "INH000014182", "khushalgrowwealth@gmail.com", "TGI"),
    ("66576c6adef5af0013c59a82", "CHIRANJEEVI KUMAR", "INH000028316", "Chiranjeevi201@gmail.com", "TGI"),
    ("666711c36a9ba40013cec685", "BOTHIKAR ADITYA", "INH000027399", "aditya.bothikar@gmail.com", "TGI"),
    # Page 21
    ("6734737fc2547400139de296", "CHUDASAMA JAYKUMAR M", "INH000028273", "jaychudasama008@gmail.com", "PP"),
    ("660bdf7f8060980014e02700", "TRENDING TRAINING AND TRADING PRIVATE LIMITED", "INH000027122", "ra.t3compliance@gmail.com", "TGI"),
    ("68b879c219c81700139c5f42", "NISHANT YADAV", "INH000025902", "Nishanty1996@gmail.com", "PP"),
    ("62a1ec8164b60533dd67264d", "Financial Independence Services", "INH000021058", "financialindependenceservice@gmail.com", "TGI"),
    ("63636438ba959d00387138e5", "Stockace Financial Services", "INH000010326", "stockacefinancialservices@gmail.com", "TGI"),
    ("6a337b2c9ad2dd0013b240c4", "SANDEEP RINWA", "INH000027098", "tradexa02@gmail.com", "TGI"),
    ("6731d6a02426e1001314e38d", "Madhur Suryavanshi", "INH000025878", "MadhurRa456@gmail.com", "TGI"),
    ("615afbdba0023812093170e3", "Jayesh Mahesh Jobanputra (intradaygeeks)", "INH000017125", "JayeshThakkar9096@gmail.com", "TGI"),
    ("6a4f7f007cc31d00131a9084", "ARUNIMA RAI PROPRIETOR SV STOCK RESEARCH", "INH000011909", "cosmo@svstockresearch.in", "TGI"),
    # Page 22
    ("661cb4abfac16e0013f3026d", "Manish kushwaha", "INH000015604", "manishkushwaha.value@gmail.com", "Both"),
    ("696f29d89e1b730013184b75", "INFINITY STOXRESEARCH", "INH000029148", "infinitystoxresearch@gmail.com", "PP"),
    ("66d5674c82e43a0013dbc67b", "Chetankumar Divakar Amin", "INH000028185", "chetanamin_1@hotmail.com", "TGI"),
    ("684a4787dc46d2001365ecce", "SIA RAVI AHUJA", "INH000028547", "kanchan90nagpal@gmail.com", "TGI"),
    ("6a88686c0011bc00139b8f38", "KHODA ANJANA NITESHKUMAR", "INH000028440", "khodaanjana078@gmail.com", "TGI"),
]


def build_sebi_master_files():
    data_dir = os.path.join(PROJECT_ROOT, "data")
    os.makedirs(data_dir, exist_ok=True)
    
    excel_path_root = os.path.join(PROJECT_ROOT, "sebi_master_creators.xlsx")
    excel_path_data = os.path.join(data_dir, "sebi_master_creators.xlsx")

    # Deduplicate entries by User ID
    unique_map = {}
    for uid, name, reg_no, email, prod in MASTER_ENTRIES:
        clean_uid = uid.strip().lower()
        if clean_uid not in unique_map:
            unique_map[clean_uid] = {
                "User ID": uid.strip(),
                "RA's Name": name.strip(),
                "SEBI Regn No.": reg_no.strip(),
                "Email": email.strip(),
                "Product": prod.strip()
            }
        else:
            # Combine product types if multiple
            existing_prod = unique_map[clean_uid]["Product"]
            if prod.strip() not in existing_prod:
                unique_map[clean_uid]["Product"] = f"{existing_prod}, {prod.strip()}"

    print(f"Total raw entries: {len(MASTER_ENTRIES)}")
    print(f"Unique SEBI Creator IDs: {len(unique_map)}")

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SEBI Master List"

    headers = ["User ID", "RA's Name", "SEBI Regn No.", "Email", "Product"]
    ws.append(headers)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, item in enumerate(unique_map.values(), start=2):
        row_vals = [
            item["User ID"],
            item["RA's Name"],
            item["SEBI Regn No."],
            item["Email"],
            item["Product"]
        ]
        ws.append(row_vals)
        for col_idx in range(1, len(row_vals) + 1):
            ws.cell(row=row_idx, column=col_idx).border = border_thin

    # Auto-width
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col_letter].width = max(14, max_len + 3)

    wb.save(excel_path_root)
    wb.save(excel_path_data)
    print(f"Saved SEBI master excel to {excel_path_root} and {excel_path_data}")

    # Also save as JSON for fast lookup
    import json
    json_path = os.path.join(data_dir, "sebi_master_creators.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(unique_map, f, indent=2)
    print(f"Saved SEBI master JSON to {json_path}")


if __name__ == "__main__":
    build_sebi_master_files()

"""
Telegram Integration & SEBI Verification Module
===============================================
Enhances the daily settlement/payout review pipeline by:
  1. Filtering settlements with payoutAmount >= ₹1,000.
  2. Sorting eligible settlements descending (Highest -> Lowest).
  3. Detecting Telegram integration products (vig/{product_id} or integratedGroup).
  4. Matching Creator IDs against the organization's SEBI Master Excel list.
  5. Generating separate 'SEBI Registered: Yes' and 'SEBI Registered: No' columns.
  6. Accumulating records over a 10-day audit period into reports/telegram_sebi_10day_records.json.
"""

import os
import re
import json
import datetime
from typing import Dict, List, Any, Optional, Tuple, Set
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = HERE
REPORTS_DIR = os.path.join(PROJECT_ROOT, "reports")
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
TEN_DAY_RECORDS_FILE = os.path.join(REPORTS_DIR, "telegram_sebi_10day_records.json")
TEN_DAY_EXCEL_FILE = os.path.join(REPORTS_DIR, "telegram_sebi_10day_report.xlsx")
TEN_DAY_CSV_FILE = os.path.join(REPORTS_DIR, "telegram_sebi_10day_report.csv")

# Master Excel locations
DEFAULT_EXCEL_PATHS = [
    os.path.join(PROJECT_ROOT, "sebi_master_creators.xlsx"),
    os.path.join(DATA_DIR, "sebi_master_creators.xlsx"),
    os.path.join(PROJECT_ROOT, "sebi_registered_creators.xlsx")
]

SEBI_MASTER_CACHE: Optional[Dict[str, Dict[str, Any]]] = None


def normalize_creator_id(cid: Any) -> str:
    """Normalize Creator ID: strip whitespace and lowercase for consistent hex comparison."""
    if not cid:
        return ""
    cid_str = str(cid).strip().lower()
    # MongoDB 24-char hex pattern
    m = re.search(r"([0-9a-f]{24})", cid_str)
    return m.group(1) if m else cid_str


def load_sebi_master_excel(excel_path: Optional[str] = None) -> Tuple[Dict[str, Dict[str, Any]], Optional[str]]:
    """
    Dynamically load verified SEBI registered creators from master Excel file.
    Returns:
      (lookup_map: {normalized_creator_id: record_dict}, error_message: Optional[str])
    """
    global SEBI_MASTER_CACHE
    if SEBI_MASTER_CACHE is not None and excel_path is None:
        return SEBI_MASTER_CACHE, None

    candidate_paths = [excel_path] if excel_path else DEFAULT_EXCEL_PATHS
    chosen_path = None
    for p in candidate_paths:
        if p and os.path.exists(p) and os.path.getsize(p) > 100:
            chosen_path = p
            break

    if not chosen_path:
        # Fallback to json if present
        json_fallback = os.path.join(DATA_DIR, "sebi_master_creators.json")
        if os.path.exists(json_fallback):
            try:
                with open(json_fallback, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    SEBI_MASTER_CACHE = data
                    return data, None
            except Exception as e:
                return {}, f"Failed to load fallback JSON: {e}"
        return {}, "SEBI Master Excel list not found"

    try:
        wb = openpyxl.load_workbook(chosen_path, data_only=True)
        ws = wb.active

        headers = [str(cell.value or "").strip() for cell in ws[1]]
        headers_lower = [h.lower() for h in headers]

        # Identify required User ID / Creator ID column
        id_col_idx = None
        for cand in ("user id", "creator id", "creator_id", "userid", "id"):
            if cand in headers_lower:
                id_col_idx = headers_lower.index(cand)
                break

        if id_col_idx is None:
            return {}, f"Required 'User ID' or 'Creator ID' column missing in {chosen_path}"

        name_col_idx = headers_lower.index("ra's name") if "ra's name" in headers_lower else (headers_lower.index("username") if "username" in headers_lower else None)
        reg_col_idx = None
        for cand in ("sebi regn no.", "sebi reg no", "sebi registration number", "reg no", "registration number"):
            if cand in headers_lower:
                reg_col_idx = headers_lower.index(cand)
                break

        email_col_idx = headers_lower.index("email") if "email" in headers_lower else None
        prod_col_idx = headers_lower.index("product") if "product" in headers_lower else None

        master_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= id_col_idx:
                continue
            raw_id = row[id_col_idx]
            norm_id = normalize_creator_id(raw_id)
            if not norm_id:
                continue

            record = {
                "creator_id": norm_id,
                "ra_name": str(row[name_col_idx] if name_col_idx is not None and name_col_idx < len(row) else "").strip(),
                "sebi_reg_no": str(row[reg_col_idx] if reg_col_idx is not None and reg_col_idx < len(row) else "").strip() or "N/A",
                "email": str(row[email_col_idx] if email_col_idx is not None and email_col_idx < len(row) else "").strip(),
                "product_type": str(row[prod_col_idx] if prod_col_idx is not None and prod_col_idx < len(row) else "").strip()
            }
            master_map[norm_id] = record

        SEBI_MASTER_CACHE = master_map
        return master_map, None
    except Exception as e:
        return {}, f"Error reading SEBI Master Excel: {str(e)}"


CREATOR_PRODUCTS_CACHE_FILE = os.path.join(REPORTS_DIR, "creator_products_cache.json")


def fetch_creator_products_from_api(creator_id: str, token: str) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """
    Fetch sold products for a creator using the official internal dashboard getCreatorKundli endpoint.
    """
    import urllib.request
    clean_id = normalize_creator_id(creator_id)
    if not clean_id or not token:
        return [], "Missing creatorId or token"

    url = f"https://prod.api.cosmofeed.com/api/internal_dashboard/getCreatorKundli?type=userId&value={clean_id}&requestedAction=lastHundredSoldProducts"
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {token}",
        "User-Agent": "CosmofeedPayoutAudit/2.0",
        "Accept": "application/json"
    })
    try:
        with urllib.request.urlopen(req, timeout=12) as resp:
            d = json.loads(resp.read().decode("utf-8"))
            if not isinstance(d, dict):
                return [], "Invalid API response structure"
            data_val = d.get("data")
            prods = []
            if isinstance(data_val, list):
                prods = data_val
            elif isinstance(data_val, dict):
                inner_val = data_val.get("data")
                if isinstance(inner_val, list):
                    prods = inner_val
                elif isinstance(inner_val, dict):
                    res = inner_val.get("lastHundredSoldProducts")
                    if isinstance(res, list):
                        prods = res
                else:
                    res = data_val.get("lastHundredSoldProducts")
                    if isinstance(res, list):
                        prods = res
            return [p for p in prods if isinstance(p, dict)], None
    except Exception as e:
        return [], str(e)


def populate_creator_products_cache(creator_ids: List[str],
                                    token: Optional[str] = None,
                                    max_workers: int = 12) -> Dict[str, List[Dict[str, Any]]]:
    """
    Populate and persist product list for creator IDs so subsequent runs are instant.
    """
    os.makedirs(REPORTS_DIR, exist_ok=True)
    cache: Dict[str, List[Dict[str, Any]]] = {}
    if os.path.exists(CREATOR_PRODUCTS_CACHE_FILE):
        try:
            with open(CREATOR_PRODUCTS_CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
        except Exception:
            cache = {}

    missing_ids = [cid for cid in creator_ids if cid and cid not in cache]
    if not missing_ids or not token:
        return cache

    import concurrent.futures
    def _fetch(cid):
        prods, err = fetch_creator_products_from_api(cid, token)
        return cid, prods, err

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        results = list(ex.map(_fetch, missing_ids))

    updated = False
    for cid, prods, err in results:
        # Cache even if empty to avoid hammering API on creators with 0 products
        if not err or prods:
            cache[cid] = prods
            updated = True

    if updated:
        try:
            with open(CREATOR_PRODUCTS_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(cache, f, indent=2)
        except Exception:
            pass

    return cache


def is_telegram_product(product_url: str = "", product_type: str = "", product_id: str = "") -> Tuple[bool, str, str]:
    """
    Detect whether a product uses Telegram integration (vig/{product_id} or integratedGroup).
    Returns:
      (is_telegram: bool, clean_product_id: str, product_link: str)
    """
    url_str = str(product_url or "").strip()
    ptype_str = str(product_type or "").strip().lower()
    pid_str = str(product_id or "").strip()

    # Match vig/{product_id} pattern
    vig_match = re.search(r"vig/([0-9a-fA-F]{24}|[0-9a-zA-Z_-]+)", url_str)
    if vig_match or "/vig/" in url_str.lower() or ptype_str in ("vig", "integratedgroup"):
        clean_pid = vig_match.group(1) if vig_match else pid_str
        constructed_link = url_str or (f"https://superprofile.bio/vig/{clean_pid}" if clean_pid else "")
        return True, clean_pid or pid_str, constructed_link

    # Check for general product ID
    clean_pid = pid_str
    if not clean_pid and url_str:
        m = re.search(r"([0-9a-fA-F]{24})", url_str)
        if m:
            clean_pid = m.group(1)

    return False, clean_pid, url_str


def verify_single_settlement(row: Dict[str, Any],
                             sebi_master: Dict[str, Dict[str, Any]],
                             sebi_load_error: Optional[str] = None,
                             creator_products_map: Optional[Dict[str, List[Dict[str, Any]]]] = None) -> Dict[str, Any]:
    """
    Process a single daily settlement row through the Telegram + SEBI verification engine.
    Rules:
      - Only settlements with payoutAmount >= 1000 are eligible.
      - Telegram integration is detected via vig/{product_id}, integratedGroup, or creator sold products.
      - Non-Telegram creators are not flagged (Not Applicable).
      - Telegram creators are matched against SEBI master list.
      - Separate 'sebi_registered_yes' and 'sebi_registered_no' columns.
    """
    res = dict(row)

    # 1. Parse settlement amount
    try:
        amt = float(row.get("payoutAmount") or row.get("totalMemoAmount") or 0.0)
    except (ValueError, TypeError):
        amt = 0.0
    res["payoutAmount"] = amt

    # 2. Check ₹1,000 threshold
    if amt < 1000.0:
        res["telegramIntegration"] = False
        res["telegramEligible"] = False
        res["sebiRegisteredYes"] = "—"
        res["sebiRegisteredNo"] = "—"
        res["sebiVerificationStatus"] = "Not Applicable"
        res["sebiReviewStatus"] = "Normal"
        res["sebiRegistrationNumber"] = "N/A"
        return res

    res["telegramEligible"] = True

    # 3. Detect Telegram integration
    cid = normalize_creator_id(row.get("creatorId") or row.get("userId") or row.get("creator_id"))

    # Direct product fields on settlement row (if present)
    prod_url = row.get("productLink") or row.get("productUrl") or row.get("url") or ""
    prod_type = row.get("productType") or row.get("product_type") or ""
    prod_id = row.get("productId") or row.get("product_id") or ""
    prod_title = row.get("productName") or row.get("productTitle") or row.get("title") or ""

    is_tele, clean_pid, clean_link = is_telegram_product(product_url=prod_url, product_type=prod_type, product_id=prod_id)

    # Deep product check: inspect creator's sold products from cache / API
    creator_prods = []
    if creator_products_map and cid in creator_products_map:
        creator_prods = creator_products_map[cid]
    elif row.get("allProducts") or row.get("products"):
        creator_prods = row.get("allProducts") or row.get("products") or []

    if not is_tele and creator_prods:
        for p in creator_prods:
            if not isinstance(p, dict):
                continue
            p_url = str(p.get("productLink") or p.get("url") or p.get("productUrl") or "")
            p_type = str(p.get("productType") or "")
            p_id = str(p.get("_id") or p.get("productId") or "")
            p_title = str(p.get("productTtile") or p.get("productTitle") or p.get("title") or "")
            p_is_tele, p_clean_pid, p_clean_link = is_telegram_product(product_url=p_url, product_type=p_type, product_id=p_id)
            if p_is_tele:
                is_tele = True
                clean_pid = p_clean_pid
                clean_link = p_clean_link
                prod_title = p_title
                break

    master_match = sebi_master.get(cid)
    match_type = "Primary Creator ID" if master_match else "None"

    # Check connected Creator IDs if not matched directly
    if not master_match:
        conn_ids = row.get("connectedCreatorIds") or row.get("connectedIds") or []
        if isinstance(conn_ids, str):
            conn_ids = [c.strip() for c in conn_ids.split(",") if c.strip()]
        for conn_id in conn_ids:
            norm_conn = normalize_creator_id(conn_id)
            if norm_conn in sebi_master:
                master_match = sebi_master[norm_conn]
                match_type = f"Connected Creator ID ({conn_id})"
                break

    # If product not directly marked, but master list specifies TGI (Telegram Integration) or Both
    if not is_tele and not prod_url and not prod_type and master_match:
        m_prod = master_match.get("product_type", "").upper()
        if "TGI" in m_prod or "BOTH" in m_prod:
            is_tele = True
            if not clean_link and clean_pid:
                clean_link = f"https://superprofile.bio/vig/{clean_pid}"

    res["telegramIntegration"] = is_tele
    res["telegramProductId"] = clean_pid
    res["telegramProductLink"] = clean_link
    res["telegramProductName"] = prod_title or ("Telegram Channel / Group" if is_tele else (row.get("productName") or "Digital Product"))

    # 4. If Non-Telegram, do not flag through this module
    if not is_tele:
        res["sebiRegisteredYes"] = "—"
        res["sebiRegisteredNo"] = "—"
        res["sebiVerificationStatus"] = "Not Applicable"
        res["sebiReviewStatus"] = "Normal"
        res["sebiRegistrationNumber"] = "N/A"
        return res

    # 5. Handle SEBI master list availability error
    if sebi_load_error:
        res["sebiRegisteredYes"] = "—"
        res["sebiRegisteredNo"] = "—"
        res["sebiVerificationStatus"] = "Unable to Validate"
        res["sebiReviewStatus"] = "Manual Review Required"
        res["sebiRegistrationNumber"] = "N/A"
        res["sebiErrorReason"] = f"Master list unavailable: {sebi_load_error}"
        return res

    # 6. Apply SEBI verification status
    res["sebiMatchType"] = match_type

    if master_match:
        # Verified SEBI Creator using Telegram
        res["sebiRegisteredYes"] = "Yes"
        res["sebiRegisteredNo"] = "—"
        res["sebiVerificationStatus"] = "Verified"
        res["sebiRegistrationNumber"] = master_match.get("sebi_reg_no") or "N/A"
        res["sebiReviewStatus"] = "Normal"
    else:
        # Using Telegram Integration but NOT found in SEBI master list: FLAGGED FOR MANUAL REVIEW!
        res["sebiRegisteredYes"] = "—"
        res["sebiRegisteredNo"] = "No"
        res["sebiVerificationStatus"] = "Not Verified"
        res["sebiRegistrationNumber"] = "N/A"
        res["sebiReviewStatus"] = "Manual Review Required"
        res["sebiReviewReason"] = "Creator uses Telegram integration (vig/) product with payout >= ₹1,000 but is NOT found in the verified SEBI-registered list."

    return res


def verify_all_settlements(settlements: List[Dict[str, Any]],
                           excel_path: Optional[str] = None,
                           token: Optional[str] = None,
                           fetch_remote: bool = True,
                           workers: int = 12) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Process entire daily settlement batch:
      1. Resolves and caches products for all eligible creators (payout >= 1000).
      2. Detects Telegram integration (vig/{product_id} or integratedGroup).
      3. Verifies against SEBI master list.
      4. Accumulates 10-day tracking records.
    Returns:
      (verified_settlements: list, summary_stats: dict)
    """
    sebi_master, load_err = load_sebi_master_excel(excel_path)

    # Resolve token if not explicitly provided
    resolved_token = token or os.environ.get("COSMOFEED_TOKEN")
    if not resolved_token:
        try:
            import payout_audit_agent as _agent
            resolved_token = getattr(_agent, "DEFAULT_TOKEN", None)
        except Exception:
            resolved_token = None

    # Identify eligible creator IDs for product pre-fetching
    eligible_cids = []
    for r in settlements:
        try:
            amt = float(r.get("payoutAmount") or r.get("totalMemoAmount") or 0.0)
        except (ValueError, TypeError):
            amt = 0.0
        if amt >= 1000.0:
            cid = normalize_creator_id(r.get("creatorId") or r.get("userId") or r.get("creator_id"))
            if cid:
                eligible_cids.append(cid)

    # Pre-fetch and cache creator products
    creator_products_map = {}
    if fetch_remote and eligible_cids and resolved_token:
        creator_products_map = populate_creator_products_cache(eligible_cids, token=resolved_token, max_workers=workers)
    elif os.path.exists(CREATOR_PRODUCTS_CACHE_FILE):
        try:
            with open(CREATOR_PRODUCTS_CACHE_FILE, "r", encoding="utf-8") as f:
                creator_products_map = json.load(f)
        except Exception:
            creator_products_map = {}

    verified_rows = []
    for r in settlements:
        v = verify_single_settlement(r, sebi_master, sebi_load_error=load_err, creator_products_map=creator_products_map)
        verified_rows.append(v)

    # Sort settlements >= 1,000 in descending settlement amount order (Highest -> Lowest)
    # Maintain non-eligible records at the bottom or maintain full sort
    def get_sort_key(item):
        amt = float(item.get("payoutAmount") or 0.0)
        is_tele = 1 if item.get("telegramIntegration") else 0
        return (is_tele, amt)

    # Filter statistics
    tele_eligible = [r for r in verified_rows if r.get("telegramEligible")]
    tele_prods = [r for r in tele_eligible if r.get("telegramIntegration")]
    sebi_yes = [r for r in tele_prods if r.get("sebiRegisteredYes") == "Yes"]
    sebi_no = [r for r in tele_prods if r.get("sebiRegisteredNo") == "No"]
    manual_review = [r for r in tele_prods if r.get("sebiReviewStatus") == "Manual Review Required"]
    unable_val = [r for r in tele_prods if r.get("sebiVerificationStatus") == "Unable to Validate"]

    unique_tele_creators = {normalize_creator_id(r.get("creatorId")) for r in tele_prods if r.get("creatorId")}

    stats = {
        "totalSettlements": len(settlements),
        "eligibleCount": len(tele_eligible),
        "telegramCount": len(tele_prods),
        "uniqueTelegramCreators": len(unique_tele_creators),
        "sebiYesCount": len(sebi_yes),
        "sebiNoCount": len(sebi_no),
        "manualReviewCount": len(manual_review),
        "unableToValidateCount": len(unable_val),
        "masterListSize": len(sebi_master),
        "masterListError": load_err
    }

    # Update cumulative 10-day audit records
    update_10day_audit_records(tele_prods)

    return verified_rows, stats


def update_10day_audit_records(telegram_settlements: List[Dict[str, Any]]) -> None:
    """
    Accumulate Telegram creators and settlements over a 10-day audit window.
    Saves to reports/telegram_sebi_10day_records.json and exports to Excel/CSV.
    """
    os.makedirs(REPORTS_DIR, exist_ok=True)
    today_str = datetime.date.today().strftime("%Y-%m-%d")

    records_map = {}
    if os.path.exists(TEN_DAY_RECORDS_FILE):
        try:
            with open(TEN_DAY_RECORDS_FILE, "r", encoding="utf-8") as f:
                records_map = json.load(f)
        except Exception:
            records_map = {}

    # Merge today's telegram settlements into records map
    for s in telegram_settlements:
        cid = normalize_creator_id(s.get("creatorId") or s.get("userId"))
        if not cid:
            continue

        amt = float(s.get("payoutAmount") or 0.0)
        date_seen = s.get("saleDate") or s.get("reviewDate") or today_str

        if cid not in records_map:
            records_map[cid] = {
                "creatorId": cid,
                "creatorName": s.get("displayName") or s.get("creatorName") or s.get("username") or "Creator",
                "username": s.get("username") or "N/A",
                "email": s.get("email") or "N/A",
                "phone": s.get("phone") or "N/A",
                "onboardedBy": s.get("onboardedBy") or "N/A",
                "onboardingVertical": s.get("vertical") or s.get("onboardingVertical") or "N/A",
                "telegramProductId": s.get("telegramProductId") or "N/A",
                "telegramProductName": s.get("telegramProductName") or "N/A",
                "telegramProductLink": s.get("telegramProductLink") or "N/A",
                "sebiRegisteredYes": s.get("sebiRegisteredYes") or "—",
                "sebiRegisteredNo": s.get("sebiRegisteredNo") or "No",
                "sebiVerificationStatus": s.get("sebiVerificationStatus") or "Not Verified",
                "sebiRegistrationNumber": s.get("sebiRegistrationNumber") or "N/A",
                "reviewStatus": s.get("sebiReviewStatus") or "Manual Review Required",
                "highestPayoutAmount": amt,
                "totalPayoutVolume": amt,
                "settlementCount": 1,
                "datesSeen": [date_seen],
                "firstSeen": date_seen,
                "lastSeen": date_seen
            }
        else:
            rec = records_map[cid]
            rec["totalPayoutVolume"] += amt
            rec["highestPayoutAmount"] = max(rec["highestPayoutAmount"], amt)
            rec["settlementCount"] += 1
            if date_seen not in rec["datesSeen"]:
                rec["datesSeen"].append(date_seen)
            rec["lastSeen"] = date_seen
            # If product link became available, update it
            if s.get("telegramProductLink") and not rec.get("telegramProductLink"):
                rec["telegramProductLink"] = s.get("telegramProductLink")
            # If SEBI verified, upgrade status
            if s.get("sebiRegisteredYes") == "Yes":
                rec["sebiRegisteredYes"] = "Yes"
                rec["sebiRegisteredNo"] = "—"
                rec["sebiVerificationStatus"] = "Verified"
                rec["reviewStatus"] = "Normal"
                if s.get("sebiRegistrationNumber"):
                    rec["sebiRegistrationNumber"] = s.get("sebiRegistrationNumber")

    # Save updated 10-day records JSON
    try:
        with open(TEN_DAY_RECORDS_FILE, "w", encoding="utf-8") as f:
            json.dump(records_map, f, indent=2)
    except Exception:
        pass

    # Generate 10-day Excel & CSV report
    export_10day_report(records_map)


def export_10day_report(records_map: Dict[str, Dict[str, Any]]) -> None:
    """Generate Excel and CSV reports for 10-day cumulative audit data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Telegram SEBI 10-Day Audit"

    headers = [
        "Creator ID",
        "Creator Name",
        "Username",
        "Email",
        "Phone",
        "Onboarded By",
        "Vertical",
        "Highest Payout (₹)",
        "Total 10d Payout (₹)",
        "Settlement Count",
        "Telegram Product ID",
        "Telegram Product Link",
        "SEBI Registered: Yes",
        "SEBI Registered: No",
        "SEBI Verification Status",
        "SEBI Registration Number",
        "Review Status",
        "Audit Dates Active",
        "First Seen",
        "Last Seen"
    ]
    ws.append(headers)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sort creators: Highest Payout -> Lowest
    sorted_records = sorted(records_map.values(), key=lambda r: -float(r.get("highestPayoutAmount", 0.0)))

    for row_idx, r in enumerate(sorted_records, start=2):
        row_vals = [
            r.get("creatorId", ""),
            r.get("creatorName", ""),
            r.get("username", ""),
            r.get("email", ""),
            r.get("phone", ""),
            r.get("onboardedBy", ""),
            r.get("onboardingVertical", ""),
            r.get("highestPayoutAmount", 0.0),
            r.get("totalPayoutVolume", 0.0),
            r.get("settlementCount", 1),
            r.get("telegramProductId", ""),
            r.get("telegramProductLink", ""),
            r.get("sebiRegisteredYes", "—"),
            r.get("sebiRegisteredNo", "No"),
            r.get("sebiVerificationStatus", "Not Verified"),
            r.get("sebiRegistrationNumber", "N/A"),
            r.get("reviewStatus", "Manual Review Required"),
            ", ".join(r.get("datesSeen", [])),
            r.get("firstSeen", ""),
            r.get("lastSeen", "")
        ]
        ws.append(row_vals)
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            if r.get("sebiRegisteredYes") == "Yes" and col_idx == 13:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif r.get("sebiRegisteredNo") == "No" and col_idx == 14:
                cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(c.value or "")) for c in col[:50])
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 3, 40))

    wb.save(TEN_DAY_EXCEL_FILE)

    # CSV Export
    import csv
    with open(TEN_DAY_CSV_FILE, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in sorted_records:
            writer.writerow([
                r.get("creatorId", ""),
                r.get("creatorName", ""),
                r.get("username", ""),
                r.get("email", ""),
                r.get("phone", ""),
                r.get("onboardedBy", ""),
                r.get("onboardingVertical", ""),
                r.get("highestPayoutAmount", 0.0),
                r.get("totalPayoutVolume", 0.0),
                r.get("settlementCount", 1),
                r.get("telegramProductId", ""),
                r.get("telegramProductLink", ""),
                r.get("sebiRegisteredYes", "—"),
                r.get("sebiRegisteredNo", "No"),
                r.get("sebiVerificationStatus", "Not Verified"),
                r.get("sebiRegistrationNumber", "N/A"),
                r.get("reviewStatus", "Manual Review Required"),
                ", ".join(r.get("datesSeen", [])),
                r.get("firstSeen", ""),
                r.get("lastSeen", "")
            ])

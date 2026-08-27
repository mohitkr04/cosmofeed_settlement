"""
Excel (.xlsx) multi-sheet and CSV export generation using openpyxl and standard csv.
"""

import os
import csv
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import CreatorProfile, ProductEvidence
from . import config


HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
EVIDENCE_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
REG_YES_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
REG_NO_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
BORDER_THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def export_to_excel(creators: List[CreatorProfile], output_path: Optional[str] = None) -> str:
    """
    Generate professional two-sheet Excel file:
      Sheet 1: SEBI Registered Creators (one unique row per Creator ID)
      Sheet 2: SEBI Evidence / Products (auditable product-level evidence)
    """
    out_file = output_path or config.OUTPUT_EXCEL
    os.makedirs(os.path.dirname(os.path.abspath(out_file)), exist_ok=True)

    wb = openpyxl.Workbook()

    # -------------------------------------------------------------
    # Sheet 1: SEBI Registered Creators
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "SEBI Registered Creators"

    headers1 = [
        "Creator ID",
        "Username",
        "Email",
        "Onboarded By",
        "Onboarding Vertical",
        "SEBI Registered",
        "SEBI Registration Number",
        "Primary Product ID",
        "Primary Product Name",
        "Primary Product Type",
        "Primary Product Link",
        "SEBI Evidence",
        "Discovery Source",
        "Connected Creator IDs",
        "Status",
        "Last Checked"
    ]

    ws1.append(headers1)
    for col_num in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Populate unique creator rows
    for row_idx, c in enumerate(creators, start=2):
        first_ev = c.product_evidence[0] if c.product_evidence else ProductEvidence()
        row_data = [
            c.creator_id,
            c.username or "N/A",
            c.email or "N/A",
            c.onboarded_by or "N/A",
            c.onboarding_vertical or "N/A",
            c.sebi_registered,
            c.sebi_registration_number or "N/A",
            first_ev.product_id or "N/A",
            first_ev.product_name or "N/A",
            first_ev.product_type or "N/A",
            first_ev.product_link or "N/A",
            c.sebi_evidence or "N/A",
            ", ".join(sorted(c.discovery_sources)) if c.discovery_sources else "N/A",
            ", ".join(sorted(c.connected_creator_ids)) if c.connected_creator_ids else "N/A",
            c.status,
            c.last_checked
        ]
        ws1.append(row_data)

        # Highlight SEBI Registered creators in soft green
        fill = REG_YES_FILL if c.sebi_registered == "YES" else REG_NO_FILL
        for col_num in range(1, len(row_data) + 1):
            cell = ws1.cell(row=row_idx, column=col_num)
            cell.border = BORDER_THIN
            if c.sebi_registered == "YES" and col_num == 6:
                cell.fill = fill
                cell.font = Font(bold=True, color="276A3C")

    # -------------------------------------------------------------
    # Sheet 2: SEBI Evidence / Products
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="SEBI Evidence - Products")
    headers2 = [
        "Creator ID",
        "Username",
        "Product ID",
        "Product Name",
        "Product Type",
        "Product Link",
        "SEBI Registration Text",
        "SEBI Registration Number",
        "Evidence Source",
        "Checked At"
    ]
    ws2.append(headers2)
    for col_num in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = EVIDENCE_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 2
    for c in creators:
        if c.product_evidence:
            for ev in c.product_evidence:
                row_data = [
                    c.creator_id,
                    c.username or "N/A",
                    ev.product_id or "N/A",
                    ev.product_name or "N/A",
                    ev.product_type or "N/A",
                    ev.product_link or "N/A",
                    ev.sebi_registration_text or "N/A",
                    ev.sebi_registration_number or "N/A",
                    ev.evidence_source or "N/A",
                    ev.checked_at
                ]
                ws2.append(row_data)
                for col_num in range(1, len(row_data) + 1):
                    ws2.cell(row=row_idx, column=col_num).border = BORDER_THIN
                row_idx += 1

    # Auto-adjust column widths for readability
    for ws in (ws1, ws2):
        ws.row_dimensions[1].height = 28
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col[:150]:  # sample first 150 rows for width
                val_str = str(cell.value or "")
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(12, min(max_len + 3, 50))

    wb.save(out_file)
    return out_file


def export_to_csv(creators: List[CreatorProfile], output_path: Optional[str] = None) -> str:
    """
    Generate normalized CSV file matching Sheet 1 columns.
    Ensures exactly one unique row per Creator ID.
    """
    out_file = output_path or config.OUTPUT_CSV
    os.makedirs(os.path.dirname(os.path.abspath(out_file)), exist_ok=True)

    fieldnames = [
        "Creator ID",
        "Username",
        "Email",
        "Onboarded By",
        "Onboarding Vertical",
        "SEBI Registered",
        "SEBI Registration Number",
        "Product ID",
        "Product Name",
        "Product Type",
        "Product Link",
        "SEBI Evidence",
        "Discovery Source",
        "Connected Creator IDs",
        "Status",
        "Last Checked"
    ]

    with open(out_file, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for c in creators:
            first_ev = c.product_evidence[0] if c.product_evidence else ProductEvidence()
            writer.writerow({
                "Creator ID": c.creator_id,
                "Username": c.username or "N/A",
                "Email": c.email or "N/A",
                "Onboarded By": c.onboarded_by or "N/A",
                "Onboarding Vertical": c.onboarding_vertical or "N/A",
                "SEBI Registered": c.sebi_registered,
                "SEBI Registration Number": c.sebi_registration_number or "N/A",
                "Product ID": first_ev.product_id or "N/A",
                "Product Name": first_ev.product_name or "N/A",
                "Product Type": first_ev.product_type or "N/A",
                "Product Link": first_ev.product_link or "N/A",
                "SEBI Evidence": c.sebi_evidence or "N/A",
                "Discovery Source": ", ".join(sorted(c.discovery_sources)) if c.discovery_sources else "N/A",
                "Connected Creator IDs": ", ".join(sorted(c.connected_creator_ids)) if c.connected_creator_ids else "N/A",
                "Status": c.status,
                "Last Checked": c.last_checked
            })

    return out_file
